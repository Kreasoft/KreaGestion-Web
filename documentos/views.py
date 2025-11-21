from utilidades.utils import clean_id
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import DocumentoCompra, ItemDocumentoCompra, HistorialPagoDocumento
from .forms import (
    DocumentoCompraForm, ItemDocumentoCompraFormSet, HistorialPagoDocumentoForm, 
    BusquedaDocumentoForm
)

from empresas.models import Empresa
from articulos.models import Articulo
from core.decorators import requiere_empresa


@login_required
@requiere_empresa
@permission_required('documentos.view_documento', raise_exception=True)
def documento_compra_list(request):
    """Lista de documentos de compra"""
    # Obtener la empresa del usuario
    if request.user.is_superuser:
        # Para superusuarios, usar empresa de sesión o Kreasoft por defecto
        empresa_id = request.session.get('empresa_activa')
        if empresa_id:
            try:
                empresa = Empresa.objects.get(id=empresa_id)
            except Empresa.DoesNotExist:
                empresa = Empresa.objects.filter(nombre__icontains='Kreasoft').first()
        else:
            empresa = Empresa.objects.filter(nombre__icontains='Kreasoft').first()
        
        if not empresa:
            empresa = Empresa.objects.first()
        
        if not empresa:
            messages.error(request, 'No hay empresas configuradas en el sistema.')
            return redirect('dashboard')
            
        # Guardar empresa en sesión
        request.session['empresa_activa'] = empresa.id
    else:
        try:
            empresa = request.user.perfil.empresa
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    # Formulario de búsqueda
    search_form = BusquedaDocumentoForm(request.GET, empresa=empresa)
    
    # Obtener documentos de compra
    documentos = DocumentoCompra.objects.filter(empresa=empresa)
    
    # Aplicar filtros
    if search_form.is_valid():
        search = search_form.cleaned_data.get('search')
        tipo_documento = search_form.cleaned_data.get('tipo_documento')
        estado_documento = search_form.cleaned_data.get('estado_documento')
        estado_pago = search_form.cleaned_data.get('estado_pago')
        fecha_desde = search_form.cleaned_data.get('fecha_desde')
        fecha_hasta = search_form.cleaned_data.get('fecha_hasta')
        proveedor = search_form.cleaned_data.get('proveedor')
        
        if search:
            documentos = documentos.filter(
                Q(numero_documento__icontains=search) |
                Q(proveedor__nombre__icontains=search) |
                Q(rut_proveedor__icontains=search) |
                Q(observaciones__icontains=search)
            ).distinct()
        
        if tipo_documento:
            documentos = documentos.filter(tipo_documento=tipo_documento)
        
        if estado_documento:
            documentos = documentos.filter(estado_documento=estado_documento)
        
        if estado_pago:
            documentos = documentos.filter(estado_pago=estado_pago)
        
        if fecha_desde:
            documentos = documentos.filter(fecha_emision__gte=fecha_desde)
        
        if fecha_hasta:
            documentos = documentos.filter(fecha_emision__lte=fecha_hasta)
        
        if proveedor:
            documentos = documentos.filter(proveedor=proveedor)
    
    # Estadísticas
    stats = {
        'total_documentos': documentos.count(),
        'total_monto': documentos.aggregate(total=Sum('total_documento'))['total'] or 0,
        'documentos_pendientes': documentos.filter(estado_documento='pendiente').count(),
        'documentos_aprobados': documentos.filter(estado_documento='aprobado').count(),
        'documentos_credito': documentos.filter(estado_pago='credito').count(),
        'documentos_pagados': documentos.filter(estado_pago='pagada').count(),
        'monto_pendiente': documentos.filter(estado_pago__in=['credito', 'parcial']).aggregate(total=Sum('saldo_pendiente'))['total'] or 0,
    }
    
    # Paginación
    paginator = Paginator(documentos, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'stats': stats,
        'empresa': empresa,
    }
    
    return render(request, 'documentos/documento_compra_list.html', context)


@login_required
@requiere_empresa
@permission_required('documentos.add_documento', raise_exception=True)
def documento_compra_create(request):
    """Crear nuevo documento de compra"""
    # Obtener la empresa del usuario
    if request.user.is_superuser:
        # Para superusuarios, usar empresa de sesión o Kreasoft por defecto
        empresa_id = request.session.get('empresa_activa')
        if empresa_id:
            try:
                empresa = Empresa.objects.get(id=empresa_id)
            except Empresa.DoesNotExist:
                empresa = Empresa.objects.filter(nombre__icontains='Kreasoft').first()
        else:
            empresa = Empresa.objects.filter(nombre__icontains='Kreasoft').first()
        
        if not empresa:
            empresa = Empresa.objects.first()
        
        if not empresa:
            messages.error(request, 'No hay empresas configuradas en el sistema.')
            return redirect('dashboard')
            
        # Guardar empresa en sesión
        request.session['empresa_activa'] = empresa.id
    else:
        try:
            empresa = request.user.perfil.empresa
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    if request.method == 'POST':
        form = DocumentoCompraForm(request.POST, request.FILES, empresa=empresa)
        formset = ItemDocumentoCompraFormSet(request.POST)
        
        # Configurar el queryset de artículos para cada formulario del formset
        articulos_empresa = Articulo.objects.filter(empresa=empresa) if empresa else Articulo.objects.none()
        for form_item in formset.forms:
            form_item.fields['articulo'].queryset = articulos_empresa
        
        # Debug: verificar errores
        if not form.is_valid():
            print("ERRORES EN FORM:")
            for field, errors in form.errors.items():
                print(f"  {field}: {errors}")
            messages.error(request, f'Errores en el formulario: {form.errors}')
        
        if not formset.is_valid():
            print("ERRORES EN FORMSET:")
            for i, form_errors in enumerate(formset.errors):
                if form_errors:
                    print(f"  Form {i}: {form_errors}")
            messages.error(request, f'Errores en los items: {formset.errors}')
        
        if form.is_valid() and formset.is_valid():
            try:
                documento = form.save(commit=False)
                documento.empresa = empresa
                documento.creado_por = request.user
                
                # Vincular con Orden de Compra si se proporcionó
                orden_compra_id = clean_id(request.POST.get('orden_compra_id'))
                if orden_compra_id:
                    from compras.models import OrdenCompra
                    try:
                        orden_compra = OrdenCompra.objects.get(pk=orden_compra_id, empresa=empresa)
                        documento.orden_compra = orden_compra
                    except OrdenCompra.DoesNotExist:
                        pass
                
                documento.save()
                
                # Guardar items
                formset.instance = documento
                formset.save()
                
                # Calcular totales
                documento.calcular_totales()
                
                # Si hay OC asociada, marcarla como completada
                if documento.orden_compra:
                    documento.orden_compra.estado_orden = 'completada'
                    documento.orden_compra.save()
                
                messages.success(request, f'Documento {documento.get_tipo_documento_display()} {documento.numero_documento} creado exitosamente.')
                return redirect('documentos:documento_compra_list')
            except Exception as e:
                print(f"ERROR AL GUARDAR: {e}")
                messages.error(request, f'Error al guardar el documento: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = DocumentoCompraForm(empresa=empresa)
        formset = ItemDocumentoCompraFormSet()
        
        # Configurar el queryset de artículos para cada formulario del formset
        articulos_empresa = Articulo.objects.filter(empresa=empresa) if empresa else Articulo.objects.none()
        for form_item in formset.forms:
            form_item.fields['articulo'].queryset = articulos_empresa
    
    context = {
        'form': form,
        'formset': formset,
        'empresa': empresa,
        'titulo': 'Crear Documento de Compra',
        'articulos_empresa': Articulo.objects.filter(empresa=empresa) if empresa else Articulo.objects.none(),
    }
    
    return render(request, 'documentos/documento_compra_form.html', context)


@login_required
def documento_compra_detail(request, pk):
    """Detalle de documento de compra"""
    documento = get_object_or_404(DocumentoCompra, pk=pk)
    
    # Para superusuarios, permitir ver documentos de cualquier empresa
    # Para usuarios normales, verificar que tenga acceso a la empresa del documento
    if not request.user.is_superuser:
        try:
            if request.user.perfil.empresa != documento.empresa:
                messages.error(request, 'No tienes permisos para ver este documento.')
                return redirect('documentos:documento_compra_list')
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    # Historial de pagos con formas de pago
    historial_pagos = documento.historial_pagos.prefetch_related('formas_pago__forma_pago').all()
    
    context = {
        'documento': documento,
        'historial_pagos': historial_pagos,
    }
    
    return render(request, 'documentos/documento_compra_detail.html', context)


@login_required
@permission_required('documentos.change_documento', raise_exception=True)
def documento_compra_update(request, pk):
    """Editar documento de compra"""
    documento = get_object_or_404(DocumentoCompra, pk=pk)
    
    # Para superusuarios, permitir editar documentos de cualquier empresa
    # Para usuarios normales, verificar que tenga acceso a la empresa del documento
    if not request.user.is_superuser:
        try:
            if request.user.perfil.empresa != documento.empresa:
                messages.error(request, 'No tienes permisos para editar este documento.')
                return redirect('documentos:documento_compra_list')
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    # Verificar que se puede editar
    if not documento.puede_editar():
        messages.error(request, 'No se puede editar un documento que ya fue aprobado o procesado.')
        return redirect('documentos:documento_compra_detail', pk=documento.pk)
    
    if request.method == 'POST':
        form = DocumentoCompraForm(request.POST, request.FILES, instance=documento, empresa=documento.empresa)
        formset = ItemDocumentoCompraFormSet(request.POST, instance=documento)
        
        # Configurar el queryset de artículos para cada formulario del formset
        articulos_empresa = Articulo.objects.filter(empresa=documento.empresa)
        for form_item in formset.forms:
            form_item.fields['articulo'].queryset = articulos_empresa
        
        if form.is_valid() and formset.is_valid():
            documento = form.save()
            
            # Guardar items
            formset.save()
            
            # Calcular totales
            documento.calcular_totales()
            
            messages.success(request, f'Documento {documento.get_tipo_documento_display()} {documento.numero_documento} actualizado exitosamente.')
            return redirect('documentos:documento_compra_list')
    else:
        form = DocumentoCompraForm(instance=documento, empresa=documento.empresa)
        # Inicializar formset con los items existentes
        formset = ItemDocumentoCompraFormSet(instance=documento)
        # Si no hay items, agregar uno vacío
        if not formset.forms:
            formset = ItemDocumentoCompraFormSet(instance=documento, queryset=documento.items.all())
        
        # Configurar el queryset de artículos para cada formulario del formset
        articulos_empresa = Articulo.objects.filter(empresa=documento.empresa)
        for form_item in formset.forms:
            form_item.fields['articulo'].queryset = articulos_empresa
    
    context = {
        'form': form,
        'formset': formset,
        'documento': documento,
        'empresa': documento.empresa,
        'titulo': 'Editar Documento de Compra',
        'articulos_empresa': Articulo.objects.filter(empresa=documento.empresa),
    }
    
    return render(request, 'documentos/documento_compra_form.html', context)


@login_required
@permission_required('documentos.delete_documento', raise_exception=True)
def documento_compra_delete(request, pk):
    """Eliminar documento de compra"""
    documento = get_object_or_404(DocumentoCompra, pk=pk)
    
    # Para superusuarios, permitir eliminar documentos de cualquier empresa
    # Para usuarios normales, verificar que tenga acceso a la empresa del documento
    if not request.user.is_superuser:
        try:
            if request.user.perfil.empresa != documento.empresa:
                messages.error(request, 'No tienes permisos para eliminar este documento.')
                return redirect('documentos:documento_compra_list')
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    # Verificar que se puede eliminar
    if not documento.puede_editar():
        messages.error(request, 'No se puede eliminar un documento que ya fue aprobado o procesado.')
        return redirect('documentos:documento_compra_detail', pk=documento.pk)
    
    if request.method == 'POST':
        tipo_doc = documento.get_tipo_documento_display()
        numero_doc = documento.numero_documento
        documento.delete()
        messages.success(request, f'Documento {tipo_doc} {numero_doc} eliminado exitosamente.')
        return redirect('documentos:documento_compra_list')
    
    context = {
        'documento': documento,
    }
    
    return render(request, 'documentos/documento_compra_confirm_delete.html', context)


@login_required
@requiere_empresa
def documento_compra_cambiar_estado_pago(request, pk):
    """Cambiar estado de pago del documento"""
    documento = get_object_or_404(DocumentoCompra, pk=pk)
    
    # Verificar que el usuario tenga acceso a la empresa del documento
    if not request.user.is_superuser:
        try:
            if request.user.perfil.empresa != documento.empresa:
                messages.error(request, 'No tienes permisos para modificar este documento.')
                return redirect('documentos:documento_compra_list')
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado_pago')
        if nuevo_estado in ['pendiente', 'credito', 'pagada']:
            documento.estado_pago = nuevo_estado
            documento.save()
            
            # Si cambia a crédito, registrar en cuenta corriente
            if nuevo_estado == 'credito' and documento.estado_documento == 'activo':
                documento.registrar_en_cuenta_corriente()
                messages.success(request, f'Documento {documento.get_tipo_documento_display()} {documento.numero_documento} cambiado a Crédito y registrado en cuenta corriente.')
            else:
                messages.success(request, f'Estado de pago del documento {documento.get_tipo_documento_display()} {documento.numero_documento} actualizado.')
        else:
            messages.error(request, 'Estado de pago inválido.')
    
    return redirect('documentos:documento_compra_detail', pk=documento.pk)


@login_required
def dashboard_documentos(request):
    """Dashboard de documentos de compra"""
    # Obtener la empresa del usuario
    if request.user.is_superuser:
        empresa = Empresa.objects.first()
        if not empresa:
            messages.error(request, 'No hay empresas configuradas en el sistema.')
            return redirect('dashboard')
    else:
        try:
            empresa = request.user.perfil.empresa
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    # Estadísticas generales
    documentos = DocumentoCompra.objects.filter(empresa=empresa)
    
    stats = {
        'total_documentos': documentos.count(),
        'total_monto': documentos.aggregate(total=Sum('total_documento'))['total'] or 0,
        'documentos_pendientes': documentos.filter(estado_documento='pendiente').count(),
        'documentos_aprobados': documentos.filter(estado_documento='aprobado').count(),
        'documentos_credito': documentos.filter(estado_pago='credito').count(),
        'documentos_pagados': documentos.filter(estado_pago='pagada').count(),
        'monto_pendiente': documentos.filter(estado_pago__in=['credito', 'parcial']).aggregate(total=Sum('saldo_pendiente'))['total'] or 0,
        'documentos_vencidos': documentos.filter(
            estado_pago__in=['credito', 'parcial'],
            fecha_vencimiento__lt=timezone.now().date()
        ).count(),
    }
    
    # Documentos recientes
    documentos_recientes = documentos.order_by('-fecha_creacion')[:10]
    
    # Documentos próximos a vencer (próximos 30 días)
    fecha_limite = timezone.now().date() + timedelta(days=30)
    documentos_por_vencer = documentos.filter(
        estado_pago__in=['credito', 'parcial'],
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gt=timezone.now().date()
    ).order_by('fecha_vencimiento')[:10]
    
    context = {
        'stats': stats,
        'documentos_recientes': documentos_recientes,
        'documentos_por_vencer': documentos_por_vencer,
        'empresa': empresa,
    }
    
    return render(request, 'documentos/dashboard_documentos.html', context)


@login_required
def get_articulo_info(request, articulo_id):
    """Obtener información de un artículo via AJAX"""
    try:
        articulo = Articulo.objects.get(pk=articulo_id)
        return JsonResponse({
            'success': True,
            'codigo': articulo.codigo,
            'nombre': articulo.nombre,
            'precio': float(articulo.precio_venta) if articulo.precio_venta else 0,
            'stock': float(articulo.stock_actual) if articulo.stock_actual else 0,
        })
    except Articulo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Artículo no encontrado'})


@login_required
@requiere_empresa
def buscar_proveedor_por_rut(request):
    """Buscar proveedor por RUT via AJAX"""
    from proveedores.models import Proveedor
    
    rut = request.GET.get('rut', '').strip()
    
    if not rut:
        return JsonResponse({'success': False, 'message': 'RUT no proporcionado'})
    
    # Obtener empresa del usuario
    if request.user.is_superuser:
        empresa_id = request.session.get('empresa_activa')
        empresa = Empresa.objects.filter(id=empresa_id).first()
    else:
        try:
            empresa = request.user.perfil.empresa
        except:
            return JsonResponse({'success': False, 'message': 'Usuario sin empresa asociada'})
    
    try:
        proveedor = Proveedor.objects.get(rut=rut, empresa=empresa)
        return JsonResponse({
            'success': True,
            'id': proveedor.id,
            'nombre': proveedor.nombre,
            'rut': proveedor.rut,
            'direccion': proveedor.direccion or '',
            'telefono': proveedor.telefono or '',
            'email': proveedor.email or ''
        })
    except Proveedor.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'No se encontró proveedor con ese RUT'
        })


@login_required
@requiere_empresa
def ordenes_compra_disponibles(request):
    """Listar órdenes de compra disponibles para un proveedor via AJAX"""
    from compras.models import OrdenCompra
    
    proveedor_id = clean_id(request.GET.get('proveedor_id'))
    
    if not proveedor_id:
        return JsonResponse({'success': False, 'message': 'ID de proveedor no proporcionado'})
    
    # Obtener empresa del usuario
    if request.user.is_superuser:
        empresa_id = request.session.get('empresa_activa')
        empresa = Empresa.objects.filter(id=empresa_id).first()
    else:
        try:
            empresa = request.user.perfil.empresa
        except:
            return JsonResponse({'success': False, 'message': 'Usuario sin empresa asociada'})
    
    # Buscar OCs del proveedor que NO estén completadas
    ordenes = OrdenCompra.objects.filter(
        proveedor_id=proveedor_id,
        empresa=empresa,
        estado_orden__in=['en_proceso', 'aprobada']  # Solo no completadas
    ).order_by('-fecha_orden')
    
    ordenes_data = []
    for oc in ordenes:
        ordenes_data.append({
            'id': oc.id,
            'numero_orden': oc.numero_orden,
            'fecha_orden': oc.fecha_orden.strftime('%d/%m/%Y'),
            'total_orden': oc.total_orden,
            'items_count': oc.items.count(),
            'estado': oc.get_estado_orden_display(),
            'bodega_id': oc.bodega_id if oc.bodega else None,
            'bodega_nombre': oc.bodega.nombre if oc.bodega else ''
        })
    
    return JsonResponse({
        'success': True,
        'ordenes': ordenes_data,
        'count': len(ordenes_data)
    })


@login_required
@requiere_empresa
def orden_compra_detalle_json(request, pk):
    """Obtener detalle completo de una orden de compra via AJAX"""
    from compras.models import OrdenCompra
    
    # Obtener empresa del usuario
    if request.user.is_superuser:
        empresa_id = request.session.get('empresa_activa')
        empresa = Empresa.objects.filter(id=empresa_id).first()
    else:
        try:
            empresa = request.user.perfil.empresa
        except:
            return JsonResponse({'success': False, 'message': 'Usuario sin empresa asociada'})
    
    try:
        orden = OrdenCompra.objects.get(pk=pk, empresa=empresa)
        
        items_data = []
        for item in orden.items.all():
            items_data.append({
                'articulo_id': item.articulo_id,
                'articulo_codigo': item.articulo.codigo,
                'articulo_nombre': item.articulo.nombre,
                'cantidad_solicitada': item.cantidad_solicitada,
                'precio_unitario': item.precio_unitario,
                'descuento_porcentaje': item.descuento_porcentaje if item.descuento_porcentaje else 0,
                'impuesto_porcentaje': item.impuesto_porcentaje if item.impuesto_porcentaje else 19,
                'subtotal': item.get_subtotal()
            })
        
        return JsonResponse({
            'success': True,
            'orden': {
                'id': orden.id,
                'numero_orden': orden.numero_orden,
                'proveedor_id': orden.proveedor_id,
                'proveedor_nombre': orden.proveedor.nombre,
                'bodega_id': orden.bodega_id if orden.bodega else None,
                'bodega_nombre': orden.bodega.nombre if orden.bodega else '',
                'fecha_orden': orden.fecha_orden.strftime('%d/%m/%Y'),
                'subtotal': orden.subtotal,
                'descuentos_totales': orden.descuentos_totales,
                'impuestos_totales': orden.impuestos_totales,
                'total_orden': orden.total_orden,
                'observaciones': orden.observaciones or ''
            },
            'items': items_data
        })
    except OrdenCompra.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Orden de compra no encontrada'
        })


@login_required
@requiere_empresa
@permission_required('documentos.view_documento', raise_exception=True)
def documento_compra_export_excel(request):
    """Exportar documentos de compra a Excel (Libro de Compras)"""
    # Obtener la empresa del usuario
    if request.user.is_superuser:
        empresa_id = request.session.get('empresa_activa')
        if empresa_id:
            try:
                empresa = Empresa.objects.get(id=empresa_id)
            except Empresa.DoesNotExist:
                empresa = Empresa.objects.filter(nombre__icontains='Kreasoft').first()
        else:
            empresa = Empresa.objects.filter(nombre__icontains='Kreasoft').first()
    else:
        try:
            empresa = request.user.perfil.empresa
        except:
            messages.error(request, 'Usuario no tiene empresa asociada.')
            return redirect('dashboard')
    
    # Obtener documentos con los mismos filtros del listado
    documentos = DocumentoCompra.objects.filter(empresa=empresa).select_related(
        'proveedor', 'bodega', 'orden_compra'
    ).order_by('-fecha_emision')
    
    # Aplicar filtros de búsqueda
    search_form = BusquedaDocumentoForm(request.GET)
    if search_form.is_valid():
        search = search_form.cleaned_data.get('search')
        tipo_documento = search_form.cleaned_data.get('tipo_documento')
        estado_documento = search_form.cleaned_data.get('estado_documento')
        estado_pago = search_form.cleaned_data.get('estado_pago')
        proveedor = search_form.cleaned_data.get('proveedor')
        
        if search:
            documentos = documentos.filter(
                Q(numero_documento__icontains=search) |
                Q(proveedor__nombre__icontains=search) |
                Q(proveedor__rut__icontains=search)
            )
        if tipo_documento:
            documentos = documentos.filter(tipo_documento=tipo_documento)
        if estado_documento:
            documentos = documentos.filter(estado_documento=estado_documento)
        if estado_pago:
            documentos = documentos.filter(estado_pago=estado_pago)
        if proveedor:
            documentos = documentos.filter(proveedor=proveedor)
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro de Compras"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"LIBRO DE COMPRAS - {empresa.nombre.upper()}"
    title_cell.font = Font(bold=True, size=14, color="8B7355")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Fecha de generación
    ws.merge_cells('A2:K2')
    date_cell = ws['A2']
    date_cell.value = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    date_cell.alignment = Alignment(horizontal='center')
    date_cell.font = Font(size=10, italic=True)
    
    # Headers
    headers = [
        'Tipo Doc.', 'N° Documento', 'Fecha Emisión', 'Fecha Venc.', 
        'Proveedor', 'RUT', 'Bodega', 'Neto', 'IVA', 'Total', 'Estado'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row_num = 5
    total_neto = 0
    total_iva = 0
    total_general = 0
    
    for doc in documentos:
        ws.cell(row=row_num, column=1, value=doc.get_tipo_documento_display()).border = border
        ws.cell(row=row_num, column=2, value=doc.numero_documento).border = border
        ws.cell(row=row_num, column=3, value=doc.fecha_emision.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row_num, column=4, value=doc.fecha_vencimiento.strftime('%d/%m/%Y') if doc.fecha_vencimiento else '').border = border
        ws.cell(row=row_num, column=5, value=doc.proveedor.nombre).border = border
        ws.cell(row=row_num, column=6, value=doc.proveedor.rut).border = border
        ws.cell(row=row_num, column=7, value=doc.bodega.nombre if doc.bodega else '').border = border
        
        # Calcular neto e IVA
        neto = float(doc.subtotal or 0)
        iva = float(doc.impuestos_totales or 0)
        total = float(doc.total_documento or 0)
        
        cell_neto = ws.cell(row=row_num, column=8, value=neto)
        cell_neto.number_format = '#,##0'
        cell_neto.border = border
        cell_neto.alignment = Alignment(horizontal='right')
        
        cell_iva = ws.cell(row=row_num, column=9, value=iva)
        cell_iva.number_format = '#,##0'
        cell_iva.border = border
        cell_iva.alignment = Alignment(horizontal='right')
        
        cell_total = ws.cell(row=row_num, column=10, value=total)
        cell_total.number_format = '#,##0'
        cell_total.border = border
        cell_total.alignment = Alignment(horizontal='right')
        cell_total.font = Font(bold=True)
        
        ws.cell(row=row_num, column=11, value=doc.get_estado_documento_display()).border = border
        
        total_neto += neto
        total_iva += iva
        total_general += total
        
        row_num += 1
    
    # Totales
    row_num += 1
    ws.cell(row=row_num, column=7, value="TOTALES:").font = Font(bold=True)
    ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='right')
    
    cell_total_neto = ws.cell(row=row_num, column=8, value=total_neto)
    cell_total_neto.number_format = '#,##0'
    cell_total_neto.font = Font(bold=True)
    cell_total_neto.fill = PatternFill(start_color="F5F1E8", end_color="F5F1E8", fill_type="solid")
    cell_total_neto.alignment = Alignment(horizontal='right')
    
    cell_total_iva = ws.cell(row=row_num, column=9, value=total_iva)
    cell_total_iva.number_format = '#,##0'
    cell_total_iva.font = Font(bold=True)
    cell_total_iva.fill = PatternFill(start_color="F5F1E8", end_color="F5F1E8", fill_type="solid")
    cell_total_iva.alignment = Alignment(horizontal='right')
    
    cell_total_general = ws.cell(row=row_num, column=10, value=total_general)
    cell_total_general.number_format = '#,##0'
    cell_total_general.font = Font(bold=True, size=12)
    cell_total_general.fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    cell_total_general.font = Font(bold=True, color="FFFFFF")
    cell_total_general.alignment = Alignment(horizontal='right')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"libro_compras_{empresa.nombre}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response