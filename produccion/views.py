from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from inventario.models import Inventario
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO

from articulos.models import RecetaProduccion, InsumoReceta, OrdenProduccion, Articulo, StockArticulo
from empresas.decorators import requiere_empresa


# ==================== RECETAS DE PRODUCCIÓN ====================

@requiere_empresa
@login_required
def receta_list(request):
    """Lista de recetas de producción"""
    recetas = RecetaProduccion.objects.filter(empresa=request.empresa).select_related('producto_final')
    
    # Filtros
    search = request.GET.get('search', '')
    tipo_produccion = request.GET.get('tipo_produccion', '')
    activo = request.GET.get('activo', '')
    
    if search:
        recetas = recetas.filter(
            Q(codigo__icontains=search) |
            Q(nombre__icontains=search) |
            Q(producto_final__nombre__icontains=search)
        )
    
    if tipo_produccion:
        recetas = recetas.filter(producto_final__tipo_produccion=tipo_produccion)
    
    if activo:
        recetas = recetas.filter(activo=activo == 'true')
    
    context = {
        'recetas': recetas,
        'search': search,
        'tipo_produccion': tipo_produccion,
        'activo': activo,
    }
    return render(request, 'produccion/receta_list.html', context)


@requiere_empresa
@login_required
def receta_detail(request, pk):
    """Detalle de una receta de producción"""
    receta = get_object_or_404(RecetaProduccion, pk=pk, empresa=request.empresa)
    insumos = receta.insumos.select_related('articulo').all()
    
    context = {
        'receta': receta,
        'insumos': insumos,
    }
    return render(request, 'produccion/receta_detail.html', context)


@requiere_empresa
@login_required
def receta_create(request):
    """Crear receta de producción"""
    if request.method == 'POST':
        try:
            # Validar campos requeridos
            codigo = request.POST.get('codigo', '').strip()
            nombre = request.POST.get('nombre', '').strip()
            producto_final = request.POST.get('producto_final', '').strip()
            
            if not all([codigo, nombre, producto_final]):
                messages.error(request, 'Todos los campos obligatorios deben estar completos.')
                return redirect('produccion:receta_create')
            
            # Debug: ver valor exacto
            print(f"DEBUG CREATE - producto_final original: {repr(producto_final)}")
            
            # Extraer SOLO dígitos usando regex
            import re
            producto_final_digits = re.sub(r'\D', '', producto_final)
            print(f"DEBUG CREATE - producto_final limpio: {repr(producto_final_digits)}")
            
            if not producto_final_digits or not producto_final_digits.isdigit():
                messages.error(request, f'ID de producto inválido. Original: {repr(producto_final)}, Limpio: {repr(producto_final_digits)}')
                return redirect('produccion:receta_create')
            
            # Crear receta
            receta = RecetaProduccion.objects.create(
                empresa=request.empresa,
                codigo=codigo,
                nombre=nombre,
                descripcion=request.POST.get('descripcion', '').strip(),
                producto_final_id=int(producto_final_digits),  # Usar versión limpia
                cantidad_producir=Decimal('1'),  # Valor por defecto, se define en órdenes
                merma_estimada=Decimal(0),  # Por defecto 0
                tiempo_estimado=int(0),  # Por defecto 0
                temperatura_proceso=None,  # No se usa por ahora
                activo=True  # Siempre activo por defecto
            )
            
            # Agregar insumos
            insumo_articulos = request.POST.getlist('insumo_articulo[]')
            insumo_cantidades = request.POST.getlist('insumo_cantidad[]')
            insumo_notas = request.POST.getlist('insumo_notas[]')
            
            for i, articulo_id in enumerate(insumo_articulos):
                if articulo_id:
                    InsumoReceta.objects.create(
                        receta=receta,
                        articulo_id=articulo_id,
                        cantidad=Decimal(insumo_cantidades[i]),
                        orden=0,  # Valor por defecto
                        notas=insumo_notas[i] if i < len(insumo_notas) else ''
                    )
            
            # Redirigir sin mensaje (se mostrará en la siguiente página)
            messages.success(request, 'Receta creada exitosamente.')
            return redirect('produccion:receta_detail', pk=receta.pk)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error al crear receta: {error_detail}")  # Para debug
            messages.error(request, f'Error al crear receta: {str(e)}')
            # No redirigir, volver a mostrar el formulario
            productos = Articulo.objects.filter(
                empresa=request.empresa,
                tipo_articulo='produccion',
                activo=True
            ).order_by('nombre')
            insumos = Articulo.objects.filter(
                empresa=request.empresa,
                tipo_articulo__in=['insumo', 'ambos'],
                activo=True
            ).order_by('nombre')
            import json
            insumos_json = json.dumps([
                {'id': ins.id, 'codigo': ins.codigo, 'nombre': ins.nombre}
                for ins in insumos
            ])
            context = {
                'productos': productos,
                'insumos': insumos,
                'insumos_json': insumos_json,
            }
            return render(request, 'produccion/receta_form.html', context)
    
    # Obtener productos finales (SOLO artículos de producción)
    productos = Articulo.objects.filter(
        empresa=request.empresa,
        tipo_articulo='produccion',  # Solo productos de producción
        activo=True
    ).order_by('nombre')
    
    # Obtener insumos
    insumos = Articulo.objects.filter(
        empresa=request.empresa,
        tipo_articulo__in=['insumo', 'ambos'],
        activo=True
    ).order_by('nombre')
    
    # Preparar insumos como lista para JavaScript
    import json
    insumos_json = json.dumps([
        {'id': ins.id, 'codigo': ins.codigo, 'nombre': ins.nombre}
        for ins in insumos
    ])
    
    context = {
        'productos': productos,
        'insumos': insumos,
        'insumos_json': insumos_json,
    }
    return render(request, 'produccion/receta_form.html', context)


@requiere_empresa
@login_required
def receta_update(request, pk):
    """Editar receta de producción"""
    receta = get_object_or_404(RecetaProduccion, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        try:
            # Validar campos requeridos
            codigo = request.POST.get('codigo', '').strip()
            nombre = request.POST.get('nombre', '').strip()
            producto_final = request.POST.get('producto_final', '').strip()
            
            if not all([codigo, nombre, producto_final]):
                messages.error(request, 'Todos los campos obligatorios deben estar completos.')
                return redirect('produccion:receta_update', pk=pk)
            
            # Debug: ver valor exacto
            print(f"DEBUG - producto_final original: {repr(producto_final)}")
            
            # Extraer SOLO dígitos usando regex
            import re
            producto_final_digits = re.sub(r'\D', '', producto_final)  # Eliminar todo lo que NO sea dígito
            print(f"DEBUG - producto_final limpio: {repr(producto_final_digits)}")
            
            if not producto_final_digits or not producto_final_digits.isdigit():
                messages.error(request, f'ID de producto inválido. Original: {repr(producto_final)}, Limpio: {repr(producto_final_digits)}')
                # NO eliminar insumos, solo mostrar error
                productos = Articulo.objects.filter(
                    empresa=request.empresa,
                    tipo_articulo='produccion',
                    activo=True
                ).order_by('nombre')
                insumos = Articulo.objects.filter(
                    empresa=request.empresa,
                    tipo_articulo__in=['insumo', 'ambos'],
                    activo=True
                ).order_by('nombre')
                import json
                insumos_json = json.dumps([
                    {'id': ins.id, 'codigo': ins.codigo, 'nombre': ins.nombre}
                    for ins in insumos
                ])
                context = {
                    'receta': receta,
                    'productos': productos,
                    'insumos': insumos,
                    'insumos_json': insumos_json,
                }
                return render(request, 'produccion/receta_form.html', context)
            
            # Validar insumos ANTES de eliminar los existentes
            insumo_articulos = request.POST.getlist('insumo_articulo[]')
            insumo_cantidades = request.POST.getlist('insumo_cantidad[]')
            insumo_notas = request.POST.getlist('insumo_notas[]')
            
            if not insumo_articulos or not any(insumo_articulos):
                messages.error(request, 'Debe agregar al menos un insumo.')
                productos = Articulo.objects.filter(
                    empresa=request.empresa,
                    tipo_articulo='produccion',
                    activo=True
                ).order_by('nombre')
                insumos = Articulo.objects.filter(
                    empresa=request.empresa,
                    tipo_articulo__in=['insumo', 'ambos'],
                    activo=True
                ).order_by('nombre')
                import json
                insumos_json = json.dumps([
                    {'id': ins.id, 'codigo': ins.codigo, 'nombre': ins.nombre}
                    for ins in insumos
                ])
                context = {
                    'receta': receta,
                    'productos': productos,
                    'insumos': insumos,
                    'insumos_json': insumos_json,
                }
                return render(request, 'produccion/receta_form.html', context)
            
            # TODO VALIDADO - Ahora sí actualizar
            receta.codigo = codigo
            receta.nombre = nombre
            receta.descripcion = request.POST.get('descripcion', '').strip()
            receta.producto_final_id = int(producto_final_digits)  # Usar la versión limpia
            receta.activo = request.POST.get('activo') == 'on'
            receta.save()
            
            # AHORA SÍ eliminar insumos existentes (después de validar todo)
            receta.insumos.all().delete()
            
            # Agregar nuevos insumos
            for i, articulo_id in enumerate(insumo_articulos):
                if articulo_id:
                    InsumoReceta.objects.create(
                        receta=receta,
                        articulo_id=articulo_id,
                        cantidad=Decimal(insumo_cantidades[i]),
                        orden=0,
                        notas=insumo_notas[i] if i < len(insumo_notas) else ''
                    )
            
            messages.success(request, 'Receta actualizada exitosamente.')
            return redirect('produccion:receta_detail', pk=pk)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error al actualizar receta: {error_detail}")
            messages.error(request, f'Error al actualizar receta: {str(e)}')
            # Recargar la receta original con sus insumos
            receta.refresh_from_db()
            # No redirigir, volver a mostrar el formulario con los datos
            productos = Articulo.objects.filter(
                empresa=request.empresa,
                tipo_articulo='produccion',
                activo=True
            ).order_by('nombre')
            insumos = Articulo.objects.filter(
                empresa=request.empresa,
                tipo_articulo__in=['insumo', 'ambos'],
                activo=True
            ).order_by('nombre')
            import json
            insumos_json = json.dumps([
                {'id': ins.id, 'codigo': ins.codigo, 'nombre': ins.nombre}
                for ins in insumos
            ])
            context = {
                'receta': receta,
                'productos': productos,
                'insumos': insumos,
                'insumos_json': insumos_json,
            }
            return render(request, 'produccion/receta_form.html', context)
    
    # Obtener productos finales (SOLO artículos de producción)
    productos = Articulo.objects.filter(
        empresa=request.empresa,
        tipo_articulo='produccion',  # Solo productos de producción
        activo=True
    ).order_by('nombre')
    
    insumos = Articulo.objects.filter(
        empresa=request.empresa,
        tipo_articulo__in=['insumo', 'ambos'],
        activo=True
    ).order_by('nombre')
    
    # Preparar insumos como lista para JavaScript
    import json
    insumos_json = json.dumps([
        {'id': ins.id, 'codigo': ins.codigo, 'nombre': ins.nombre}
        for ins in insumos
    ])
    
    context = {
        'receta': receta,
        'productos': productos,
        'insumos': insumos,
        'insumos_json': insumos_json,
    }
    return render(request, 'produccion/receta_form.html', context)


@requiere_empresa
@login_required
def receta_delete(request, pk):
    """Eliminar receta de producción"""
    receta = get_object_or_404(RecetaProduccion, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        receta.delete()
        messages.success(request, 'Receta eliminada exitosamente.')
        return redirect('produccion:receta_list')
    
    context = {'receta': receta}
    return render(request, 'produccion/receta_confirm_delete.html', context)


# ==================== ÓRDENES DE PRODUCCIÓN ====================

@requiere_empresa
@login_required
def orden_list(request):
    """Lista de órdenes de producción"""
    ordenes = OrdenProduccion.objects.filter(empresa=request.empresa).select_related('receta', 'sucursal')
    
    # Filtros
    search = request.GET.get('search', '')
    estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if search:
        ordenes = ordenes.filter(
            Q(numero_orden__icontains=search) |
            Q(receta__nombre__icontains=search) |
            Q(responsable__icontains=search)
        )
    
    if estado:
        ordenes = ordenes.filter(estado=estado)
    
    if fecha_desde:
        ordenes = ordenes.filter(fecha_planificada__gte=fecha_desde)
    
    if fecha_hasta:
        ordenes = ordenes.filter(fecha_planificada__lte=fecha_hasta)
    
    # Ordenar por ID descendente (más recientes primero)
    ordenes = ordenes.order_by('-id')
    
    # Obtener recetas para el modal
    recetas = RecetaProduccion.objects.filter(empresa=request.empresa, activo=True).order_by('nombre')
    
    # Obtener bodegas activas
    from bodegas.models import Bodega
    bodegas = Bodega.objects.filter(empresa=request.empresa, activa=True).order_by('nombre')
    
    context = {
        'ordenes': ordenes,
        'recetas': recetas,
        'bodegas': bodegas,
        'search': search,
        'estado': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'produccion/orden_list.html', context)


@requiere_empresa
@login_required
def orden_detail(request, pk):
    """Detalle de una orden de producción"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    context = {'orden': orden}
    return render(request, 'produccion/orden_detail.html', context)


@requiere_empresa
@login_required
def orden_detail_json(request, pk):
    """Devuelve el detalle de una orden en formato JSON para el modal"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    data = {
        'numero_orden': orden.numero_orden,
        'estado': orden.estado,
        'porcentaje_completado': float(orden.porcentaje_completado),
        'receta_nombre': orden.receta.nombre,
        'producto_nombre': orden.receta.producto_final.nombre,
        'responsable': orden.responsable,
        'lote_produccion': orden.lote_produccion,
        'fecha_planificada': orden.fecha_planificada.strftime('%d/%m/%Y') if orden.fecha_planificada else None,
        'fecha_inicio': orden.fecha_inicio.strftime('%d/%m/%Y %H:%M') if orden.fecha_inicio else None,
        'fecha_fin': orden.fecha_fin.strftime('%d/%m/%Y %H:%M') if orden.fecha_fin else None,
        'cantidad_planificada': str(orden.cantidad_planificada),
        'cantidad_producida': str(orden.cantidad_producida),
        'merma_real': str(orden.merma_real),
        'costo_total': f'${orden.costo_total:,.0f}'.replace(',', ' '),
        'observaciones': orden.observaciones if hasattr(orden, 'observaciones') else None,
    }
    
    return JsonResponse(data)


@requiere_empresa
@login_required
def orden_create(request):
    """Crear orden de producción"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            numero_orden = request.POST.get('numero_orden', '').strip()
            receta_id = request.POST.get('receta', '').strip()
            cantidad_planificada = request.POST.get('cantidad_planificada', '').strip()
            fecha_planificada = request.POST.get('fecha_planificada', '').strip()
            responsable = request.POST.get('responsable', '').strip()
            bodega_id = request.POST.get('bodega', '').strip()
            notas = request.POST.get('notas', '').strip()
            
            # Validar campos requeridos
            if not all([numero_orden, receta_id, cantidad_planificada, bodega_id]):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Todos los campos obligatorios deben estar completos (Número, Receta, Cantidad, Bodega).'
                    })
                messages.error(request, 'Todos los campos obligatorios deben estar completos.')
                recetas = RecetaProduccion.objects.filter(empresa=request.empresa, activo=True)
                context = {'recetas': recetas}
                return render(request, 'produccion/orden_form.html', context)
            
            # Obtener bodega seleccionada
            from bodegas.models import Bodega, Sucursal
            try:
                bodega = Bodega.objects.get(id=int(bodega_id), empresa=request.empresa)
                sucursal = bodega.sucursal
            except Bodega.DoesNotExist:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La bodega seleccionada no existe.'
                    })
                messages.error(request, 'La bodega seleccionada no existe.')
                recetas = RecetaProduccion.objects.filter(empresa=request.empresa, activo=True)
                context = {'recetas': recetas}
                return render(request, 'produccion/orden_form.html', context)
            
            # Crear orden
            orden = OrdenProduccion.objects.create(
                empresa=request.empresa,
                sucursal=sucursal,
                numero_orden=numero_orden,
                receta_id=int(receta_id),
                cantidad_planificada=Decimal(cantidad_planificada),
                cantidad_producida=Decimal(0),
                merma_real=Decimal(0),
                estado='pendiente',
                fecha_planificada=fecha_planificada if fecha_planificada else None,
                responsable=responsable
            )
            
            # Si es AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Orden de producción creada exitosamente.',
                    'orden_id': orden.pk
                })
            
            messages.success(request, 'Orden de producción creada exitosamente.')
            return redirect('produccion:orden_detail', pk=orden.pk)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error al crear orden: {error_detail}")
            messages.error(request, f'Error al crear orden: {str(e)}')
            recetas = RecetaProduccion.objects.filter(empresa=request.empresa, activo=True)
            context = {'recetas': recetas}
            return render(request, 'produccion/orden_form.html', context)
    
    recetas = RecetaProduccion.objects.filter(empresa=request.empresa, activo=True)
    
    context = {'recetas': recetas}
    return render(request, 'produccion/orden_form.html', context)


@requiere_empresa
@login_required
def orden_update(request, pk):
    """Editar orden de producción"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    # Solo se pueden editar órdenes pendientes
    if orden.estado != 'pendiente':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'Solo se pueden editar órdenes pendientes.'
            })
        messages.error(request, 'Solo se pueden editar órdenes pendientes.')
        return redirect('produccion:orden_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            numero_orden = request.POST.get('numero_orden', '').strip()
            receta_id = request.POST.get('receta', '').strip()
            cantidad_planificada = request.POST.get('cantidad_planificada', '').strip()
            fecha_planificada = request.POST.get('fecha_planificada', '').strip()
            responsable = request.POST.get('responsable', '').strip()
            bodega_id = request.POST.get('bodega', '').strip()
            lote_produccion = request.POST.get('lote_produccion', '').strip()
            observaciones = request.POST.get('observaciones', '').strip()
            
            # Validar campos requeridos
            if not all([numero_orden, receta_id, cantidad_planificada, bodega_id]):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Todos los campos obligatorios deben estar completos.'
                    })
                messages.error(request, 'Todos los campos obligatorios deben estar completos.')
                return redirect('produccion:orden_update', pk=pk)
            
            # Obtener bodega seleccionada
            from bodegas.models import Bodega
            try:
                bodega = Bodega.objects.get(id=int(bodega_id), empresa=request.empresa)
                sucursal = bodega.sucursal
            except Bodega.DoesNotExist:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La bodega seleccionada no existe.'
                    })
                messages.error(request, 'La bodega seleccionada no existe.')
                return redirect('produccion:orden_update', pk=pk)
            
            # Actualizar orden
            orden.numero_orden = numero_orden
            orden.receta_id = int(receta_id)
            orden.cantidad_planificada = Decimal(cantidad_planificada)
            orden.fecha_planificada = fecha_planificada if fecha_planificada else None
            orden.responsable = responsable
            orden.sucursal = sucursal
            orden.lote_produccion = lote_produccion
            orden.observaciones = observaciones
            orden.save()
            
            # Si es AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Orden actualizada exitosamente.'
                })
            
            messages.success(request, 'Orden actualizada exitosamente.')
            return redirect('produccion:orden_detail', pk=pk)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error al actualizar orden: {error_detail}")
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al actualizar orden: {str(e)}'
                })
            messages.error(request, f'Error al actualizar orden: {str(e)}')
            return redirect('produccion:orden_update', pk=pk)
    
    recetas = RecetaProduccion.objects.filter(empresa=request.empresa, activo=True)
    
    context = {
        'orden': orden,
        'recetas': recetas,
    }
    return render(request, 'produccion/orden_form.html', context)


@requiere_empresa
@login_required
def orden_delete(request, pk):
    """Eliminar orden de producción"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        orden.delete()
        messages.success(request, 'Orden eliminada exitosamente.')
        return redirect('produccion:orden_list')
    
    context = {'orden': orden}
    return render(request, 'produccion/orden_confirm_delete.html', context)


@requiere_empresa
@login_required
def orden_iniciar(request, pk):
    """Iniciar producción de una orden"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    if orden.estado != 'pendiente':
        messages.error(request, 'Solo se pueden iniciar órdenes pendientes.')
        return redirect('produccion:orden_detail', pk=pk)
    
    if request.method == 'POST':
        orden.estado = 'en_proceso'
        orden.fecha_inicio = timezone.now()
        orden.save()
        
        messages.success(request, 'Orden iniciada exitosamente.')
        return redirect('produccion:orden_detail', pk=pk)
    
    context = {'orden': orden}
    return render(request, 'produccion/orden_iniciar.html', context)


@requiere_empresa
@login_required
def orden_finalizar(request, pk):
    """Finalizar producción de una orden"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    if orden.estado != 'en_proceso':
        messages.error(request, 'Solo se pueden finalizar órdenes en proceso.')
        return redirect('produccion:orden_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            cantidad_producida = Decimal(request.POST.get('cantidad_producida', 0))
            merma_real = Decimal(request.POST.get('merma_real', 0))
            comentarios = request.POST.get('comentarios', '').strip()
            meses_garantia = request.POST.get('meses_garantia', '').strip()
            bodega_insumos_id = request.POST.get('bodega_insumos', '').strip()
            bodega_productos_id = request.POST.get('bodega_productos', '').strip()
            
            # Validar que se seleccionaron las bodegas
            if not bodega_insumos_id or not bodega_productos_id:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Debe seleccionar las bodegas de insumos y productos.'
                    })
                messages.error(request, 'Debe seleccionar las bodegas de insumos y productos.')
                return redirect('produccion:orden_detail', pk=pk)
            
            # Obtener las bodegas
            from bodegas.models import Bodega
            try:
                bodega_insumos = Bodega.objects.get(id=int(bodega_insumos_id), empresa=request.empresa)
                bodega_productos = Bodega.objects.get(id=int(bodega_productos_id), empresa=request.empresa)
            except Bodega.DoesNotExist:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Una o ambas bodegas seleccionadas no existen.'
                    })
                messages.error(request, 'Una o ambas bodegas seleccionadas no existen.')
                return redirect('produccion:orden_detail', pk=pk)
            
            # Calcular factor de producción (cuántas veces se ejecutó la receta)
            factor_produccion = orden.cantidad_planificada / orden.receta.cantidad_producir
            
            # VALIDAR STOCK DE INSUMOS ANTES DE FINALIZAR
            insumos_sin_stock = []
            for insumo in orden.receta.insumos.all():
                cantidad_necesaria = insumo.cantidad * factor_produccion
                
                # Buscar stock del insumo
                stock = StockArticulo.objects.filter(
                    articulo=insumo.articulo,
                    sucursal=orden.sucursal
                ).first()
            
            if stock:
                if isinstance(stock.cantidad_disponible, str):
                    stock_disponible = Decimal(stock.cantidad_disponible.replace(',', '.'))
                else:
                    stock_disponible = Decimal(str(stock.cantidad_disponible))
            else:
                stock_disponible = Decimal('0')
            
            if stock_disponible < cantidad_necesaria:
                insumos_sin_stock.append({
                    'nombre': insumo.articulo.nombre,
                    'necesario': float(cantidad_necesaria),
                    'disponible': float(stock_disponible),
                    'faltante': float(cantidad_necesaria - stock_disponible)
                })
            
            # Si hay insumos sin stock, solo registrar advertencia (no bloquear)
            if insumos_sin_stock:
                print("⚠️ ADVERTENCIA: Insumos con stock insuficiente:")
                for insumo in insumos_sin_stock:
                    print(f"  • {insumo['nombre']}: Necesario {insumo['necesario']}, Disponible {insumo['disponible']}, Falta {insumo['faltante']}")
            
            # Actualizar orden
            orden.estado = 'terminada'
            orden.fecha_fin = timezone.now()
            orden.cantidad_producida = cantidad_producida
            orden.merma_real = merma_real
            orden.observaciones = comentarios
            orden.meses_garantia = int(meses_garantia) if meses_garantia else None
            orden.save()
            
            # Consumir stock de insumos
            for insumo in orden.receta.insumos.all():
                cantidad_consumir = insumo.cantidad * factor_produccion
                
                # Buscar stock del insumo en la sucursal
                stock, created = StockArticulo.objects.get_or_create(
                    articulo=insumo.articulo,
                    sucursal=orden.sucursal,
                    defaults={'cantidad_disponible': '0', 'cantidad_reservada': '0'}
                )
                
                # Reducir stock disponible
                if isinstance(stock.cantidad_disponible, str):
                    stock_actual = Decimal(stock.cantidad_disponible.replace(',', '.'))
                else:
                    stock_actual = Decimal(str(stock.cantidad_disponible))
                nuevo_stock = stock_actual - cantidad_consumir
                stock.cantidad_disponible = str(nuevo_stock)
                stock.save()
                
                # Registrar movimiento en kardex (SALIDA de insumo desde bodega de insumos)
                mov_salida = Inventario.objects.create(
                    empresa=orden.empresa,
                    bodega_origen=bodega_insumos,
                    articulo=insumo.articulo,
                    tipo_movimiento='salida',
                    cantidad=cantidad_consumir,
                    precio_unitario=Decimal(insumo.articulo.precio_costo.replace(',', '.')) if insumo.articulo.precio_costo else Decimal('0'),
                    total=cantidad_consumir * (Decimal(insumo.articulo.precio_costo.replace(',', '.')) if insumo.articulo.precio_costo else Decimal('0')),
                    motivo=f'Consumo por Producción - Orden {orden.numero_orden}',
                    descripcion=f'Insumo consumido en producción de {orden.receta.producto_final.nombre}',
                    numero_documento=str(orden.numero_orden),
                    estado='confirmado',
                    fecha_movimiento=timezone.now(),
                    creado_por=request.user
                )
                print(f"✅ Movimiento SALIDA creado: ID={mov_salida.id}, Artículo={insumo.articulo.nombre}, Bodega={bodega_insumos.nombre}, Cantidad={cantidad_consumir}")
            
            # Agregar stock del producto final
            stock_producto, created = StockArticulo.objects.get_or_create(
                articulo=orden.receta.producto_final,
                sucursal=orden.sucursal,
                defaults={'cantidad_disponible': '0', 'cantidad_reservada': '0'}
            )
            
            # Aumentar stock disponible
            if isinstance(stock_producto.cantidad_disponible, str):
                stock_actual = Decimal(stock_producto.cantidad_disponible.replace(',', '.'))
            else:
                stock_actual = Decimal(str(stock_producto.cantidad_disponible))
            nuevo_stock = stock_actual + cantidad_producida
            stock_producto.cantidad_disponible = str(nuevo_stock)
            stock_producto.save()
            
            # Registrar movimiento en kardex (ENTRADA de producto final a bodega de productos)
            mov_entrada = Inventario.objects.create(
                empresa=orden.empresa,
                bodega_destino=bodega_productos,
                articulo=orden.receta.producto_final,
                tipo_movimiento='entrada',
                cantidad=cantidad_producida,
                precio_unitario=orden.receta.costo_unitario,
                total=cantidad_producida * orden.receta.costo_unitario,
                motivo=f'Producción - Orden {orden.numero_orden}',
                descripcion=f'Producto fabricado. Merma: {merma_real}. {comentarios[:100] if comentarios else ""}',
                numero_documento=str(orden.numero_orden),
                estado='confirmado',
                fecha_movimiento=timezone.now(),
                creado_por=request.user
            )
            print(f"✅ Movimiento ENTRADA creado: ID={mov_entrada.id}, Artículo={orden.receta.producto_final.nombre}, Bodega={bodega_productos.nombre}, Cantidad={cantidad_producida}")
            
            # Si es AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Orden finalizada exitosamente.'
                })
            
            messages.success(request, 'Orden finalizada exitosamente.')
            return redirect('produccion:orden_detail', pk=pk)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error al finalizar orden: {error_detail}")
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al finalizar orden: {str(e)}'
                })
            messages.error(request, f'Error al finalizar orden: {str(e)}')
            return redirect('produccion:orden_detail', pk=pk)
    
    context = {'orden': orden}
    return render(request, 'produccion/orden_finalizar.html', context)


@requiere_empresa
@login_required
def orden_cancelar(request, pk):
    """Cancelar orden de producción"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    if orden.estado == 'terminada':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'No se pueden cancelar órdenes terminadas.'
            })
        messages.error(request, 'No se pueden cancelar órdenes terminadas.')
        return redirect('produccion:orden_detail', pk=pk)
    
    if request.method == 'POST':
        orden.estado = 'cancelada'
        orden.save()
        
        # Si es AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Orden cancelada exitosamente.'
            })
        
        messages.success(request, 'Orden cancelada exitosamente.')
        return redirect('produccion:orden_list')
    
    context = {'orden': orden}
    return render(request, 'produccion/orden_cancelar.html', context)


# ==================== REPORTES ====================

@requiere_empresa
@login_required
def reportes(request):
    """Vista principal de reportes"""
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    empresa = request.empresa
    hoy = timezone.now()
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Órdenes activas (pendiente + en_proceso)
    ordenes_activas = OrdenProduccion.objects.filter(
        empresa=empresa,
        estado__in=['pendiente', 'en_proceso']
    ).count()
    
    # Completadas este mes
    completadas_mes = OrdenProduccion.objects.filter(
        empresa=empresa,
        estado='terminada',
        fecha_fin__gte=inicio_mes
    ).count()
    
    # Recetas activas
    recetas_activas = RecetaProduccion.objects.filter(
        empresa=empresa,
        activo=True
    ).count()
    
    # Costo total del mes (calculado desde las propiedades)
    ordenes_terminadas_mes = OrdenProduccion.objects.filter(
        empresa=empresa,
        estado='terminada',
        fecha_fin__gte=inicio_mes
    ).select_related('receta')
    
    costo_total_mes = sum(orden.costo_total for orden in ordenes_terminadas_mes)
    
    context = {
        'ordenes_activas': ordenes_activas,
        'completadas_mes': completadas_mes,
        'recetas_activas': recetas_activas,
        'costo_total_mes': costo_total_mes,
    }
    
    return render(request, 'produccion/reportes.html', context)


@requiere_empresa
@login_required
def reporte_produccion(request):
    """Reporte de producción con estadísticas y análisis"""
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    estado = request.GET.get('estado', '')
    
    # Filtrar órdenes
    ordenes = OrdenProduccion.objects.filter(empresa=request.empresa).select_related('receta', 'sucursal')
    
    if fecha_desde:
        ordenes = ordenes.filter(fecha_planificada__gte=fecha_desde)
    
    if fecha_hasta:
        ordenes = ordenes.filter(fecha_planificada__lte=fecha_hasta)
    
    if estado:
        ordenes = ordenes.filter(estado=estado)
    
    # Calcular estadísticas
    total_ordenes = ordenes.count()
    ordenes_terminadas = ordenes.filter(estado='terminada').count()
    
    # Totales
    total_planificado = sum(orden.cantidad_planificada for orden in ordenes)
    total_producido = sum(orden.cantidad_producida for orden in ordenes)
    total_merma = sum(orden.merma_real for orden in ordenes)
    
    # Eficiencia promedio
    if total_planificado > 0:
        eficiencia_promedio = (total_producido / total_planificado) * 100
    else:
        eficiencia_promedio = 0
    
    # Costo total
    costo_total = sum(orden.costo_total for orden in ordenes)
    
    context = {
        'ordenes': ordenes.order_by('-fecha_planificada'),
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'estado': estado,
        'total_ordenes': total_ordenes,
        'ordenes_terminadas': ordenes_terminadas,
        'total_planificado': total_planificado,
        'total_producido': total_producido,
        'total_merma': total_merma,
        'eficiencia_promedio': eficiencia_promedio,
        'costo_total': costo_total,
    }
    
    return render(request, 'produccion/reporte_produccion.html', context)


@requiere_empresa
@login_required
def reporte_costos(request):
    """Reporte de costos de producción con análisis detallado"""
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    receta_id = request.GET.get('receta', '')
    
    # Filtrar órdenes
    ordenes = OrdenProduccion.objects.filter(
        empresa=request.empresa,
        estado='terminada'  # Solo órdenes terminadas para costos reales
    ).select_related('receta', 'receta__producto_final')
    
    if fecha_desde:
        ordenes = ordenes.filter(fecha_planificada__gte=fecha_desde)
    
    if fecha_hasta:
        ordenes = ordenes.filter(fecha_planificada__lte=fecha_hasta)
    
    if receta_id:
        ordenes = ordenes.filter(receta_id=receta_id)
    
    # Calcular costos totales
    costo_total_produccion = sum(orden.costo_total for orden in ordenes)
    costo_total_insumos = sum(orden.receta.costo_total_insumos * (orden.cantidad_planificada / orden.receta.cantidad_producir) for orden in ordenes)
    
    # Costo promedio unitario
    total_unidades = sum(orden.cantidad_producida for orden in ordenes)
    if total_unidades > 0:
        costo_promedio_unitario = costo_total_produccion / total_unidades
    else:
        costo_promedio_unitario = 0
    
    # Agrupar por receta
    from collections import defaultdict
    costos_por_receta_dict = defaultdict(lambda: {
        'receta': None,
        'total_ordenes': 0,
        'costo_total_producido': Decimal('0')
    })
    
    for orden in ordenes:
        key = orden.receta.id
        if costos_por_receta_dict[key]['receta'] is None:
            costos_por_receta_dict[key]['receta'] = orden.receta
        costos_por_receta_dict[key]['total_ordenes'] += 1
        costos_por_receta_dict[key]['costo_total_producido'] += orden.costo_total
    
    costos_por_receta = list(costos_por_receta_dict.values())
    
    # Obtener todas las recetas para el filtro
    recetas = RecetaProduccion.objects.filter(empresa=request.empresa, activo=True).order_by('nombre')
    
    context = {
        'ordenes': ordenes.order_by('-fecha_planificada'),
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'receta_id': receta_id,
        'recetas': recetas,
        'costo_total_insumos': costo_total_insumos,
        'costo_total_produccion': costo_total_produccion,
        'costo_promedio_unitario': costo_promedio_unitario,
        'total_recetas': len(costos_por_receta),
        'costos_por_receta': costos_por_receta,
    }
    
    return render(request, 'produccion/reporte_costos.html', context)


@requiere_empresa
@login_required
def exportar_reporte_produccion_excel(request):
    """Exportar reporte de producción a Excel"""
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    estado = request.GET.get('estado', '')
    
    # Filtrar órdenes
    ordenes = OrdenProduccion.objects.filter(empresa=request.empresa).select_related('receta', 'sucursal')
    
    if fecha_desde:
        ordenes = ordenes.filter(fecha_planificada__gte=fecha_desde)
    
    if fecha_hasta:
        ordenes = ordenes.filter(fecha_planificada__lte=fecha_hasta)
    
    if estado:
        ordenes = ordenes.filter(estado=estado)
    
    ordenes = ordenes.order_by('-fecha_planificada')
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Producción"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = f'REPORTE DE PRODUCCIÓN - {request.empresa.nombre}'
    ws['A1'].font = Font(bold=True, size=14, color="8B7355")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Filtros aplicados
    row = 2
    if fecha_desde or fecha_hasta:
        ws.merge_cells(f'A{row}:I{row}')
        periodo = f"Período: {fecha_desde or 'Inicio'} al {fecha_hasta or 'Hoy'}"
        ws[f'A{row}'] = periodo
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
    
    # Encabezados
    row += 1
    headers = ['Orden', 'Receta', 'Fecha', 'Planificada', 'Producida', 'Merma', 'Eficiencia %', 'Estado', 'Costo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    total_planificado = 0
    total_producido = 0
    total_merma = 0
    total_costo = 0
    
    for orden in ordenes:
        row += 1
        ws.cell(row=row, column=1, value=f"OP-{orden.numero_orden}").border = border
        ws.cell(row=row, column=2, value=orden.receta.nombre).border = border
        ws.cell(row=row, column=3, value=orden.fecha_planificada.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=4, value=float(orden.cantidad_planificada)).border = border
        ws.cell(row=row, column=5, value=float(orden.cantidad_producida)).border = border
        ws.cell(row=row, column=6, value=float(orden.merma_real)).border = border
        ws.cell(row=row, column=7, value=f"{float(orden.porcentaje_completado):.0f}%").border = border
        ws.cell(row=row, column=8, value=orden.get_estado_display()).border = border
        ws.cell(row=row, column=9, value=float(orden.costo_total)).border = border
        
        total_planificado += orden.cantidad_planificada
        total_producido += orden.cantidad_producida
        total_merma += orden.merma_real
        total_costo += orden.costo_total
    
    # Totales
    row += 1
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=4, value=float(total_planificado)).font = Font(bold=True)
    ws.cell(row=row, column=4).border = border
    ws.cell(row=row, column=5, value=float(total_producido)).font = Font(bold=True)
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=6, value=float(total_merma)).font = Font(bold=True)
    ws.cell(row=row, column=6).border = border
    ws.cell(row=row, column=9, value=float(total_costo)).font = Font(bold=True)
    ws.cell(row=row, column=9).border = border
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_produccion_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@requiere_empresa
@login_required
def exportar_orden_excel(request, pk):
    """Exportar detalle de orden con insumos a Excel"""
    orden = get_object_or_404(OrdenProduccion, pk=pk, empresa=request.empresa)
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orden {orden.numero_orden}"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    info_fill = PatternFill(start_color="F5F5F0", end_color="F5F5F0", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f'ORDEN DE PRODUCCIÓN - OP-{orden.numero_orden}'
    ws['A1'].font = Font(bold=True, size=14, color="8B7355")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Información de la orden
    row = 3
    info_data = [
        ('Receta:', orden.receta.nombre),
        ('Producto Final:', orden.receta.producto_final.nombre),
        ('Cantidad a Producir:', f"{orden.cantidad_planificada} {str(orden.receta.producto_final.unidad_medida)}"),
        ('Fecha Planificada:', orden.fecha_planificada.strftime('%d/%m/%Y')),
        ('Responsable:', orden.responsable or '-'),
        ('Estado:', orden.get_estado_display()),
    ]
    
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = info_fill
        ws.cell(row=row, column=2, value=value)
        ws.merge_cells(f'B{row}:E{row}')
        row += 1
    
    # Sección de insumos
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'INSUMOS REQUERIDOS'
    ws[f'A{row}'].font = Font(bold=True, size=12, color="8B7355")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Encabezados de insumos
    row += 1
    headers = ['Insumo', 'Cantidad por Receta', 'Cantidad Total', 'Unidad', 'Stock Disponible']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Calcular factor de producción
    factor_produccion = orden.cantidad_planificada / orden.receta.cantidad_producir
    
    # Datos de insumos
    for insumo in orden.receta.insumos.all():
        row += 1
        cantidad_total = insumo.cantidad * factor_produccion
        
        # Obtener stock disponible
        stock = StockArticulo.objects.filter(
            articulo=insumo.articulo,
            sucursal=orden.sucursal
        ).first()
        
        if stock:
            # Convertir a Decimal si es string, sino usar directamente
            if isinstance(stock.cantidad_disponible, str):
                stock_disponible = Decimal(stock.cantidad_disponible.replace(',', '.'))
            else:
                stock_disponible = Decimal(str(stock.cantidad_disponible))
        else:
            stock_disponible = Decimal('0')
        
        ws.cell(row=row, column=1, value=insumo.articulo.nombre).border = border
        ws.cell(row=row, column=2, value=float(insumo.cantidad)).border = border
        ws.cell(row=row, column=3, value=float(cantidad_total)).border = border
        ws.cell(row=row, column=4, value=str(insumo.articulo.unidad_medida)).border = border
        
        # Stock con color según disponibilidad
        cell_stock = ws.cell(row=row, column=5, value=float(stock_disponible))
        cell_stock.border = border
        if stock_disponible < cantidad_total:
            cell_stock.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            cell_stock.font = Font(color="FF0000", bold=True)
        else:
            cell_stock.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            cell_stock.font = Font(color="008000", bold=True)
    
    # Nota sobre stock
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'Nota: Los insumos en ROJO tienen stock insuficiente'
    ws[f'A{row}'].font = Font(italic=True, color="FF0000")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="orden_produccion_OP-{orden.numero_orden}.xlsx"'
    
    wb.save(response)
    return response


@requiere_empresa
@login_required
def exportar_reporte_produccion_pdf(request):
    """Exportar reporte de producción a PDF"""
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    estado = request.GET.get('estado', '')
    
    # Filtrar órdenes
    ordenes = OrdenProduccion.objects.filter(empresa=request.empresa).select_related('receta', 'sucursal')
    
    if fecha_desde:
        ordenes = ordenes.filter(fecha_planificada__gte=fecha_desde)
    
    if fecha_hasta:
        ordenes = ordenes.filter(fecha_planificada__lte=fecha_hasta)
    
    if estado:
        ordenes = ordenes.filter(estado=estado)
    
    ordenes = ordenes.order_by('-fecha_planificada')
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Contenedor de elementos
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#8B7355'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Título
    title = Paragraph(f'REPORTE DE PRODUCCIÓN<br/>{request.empresa.nombre}', title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Período
    if fecha_desde or fecha_hasta:
        periodo_text = f"Período: {fecha_desde or 'Inicio'} al {fecha_hasta or 'Hoy'}"
        periodo = Paragraph(periodo_text, styles['Normal'])
        elements.append(periodo)
        elements.append(Spacer(1, 12))
    
    # Datos de la tabla
    data = [['Orden', 'Receta', 'Fecha', 'Plan.', 'Prod.', 'Merma', 'Efic.%', 'Estado']]
    
    total_planificado = 0
    total_producido = 0
    total_merma = 0
    
    for orden in ordenes:
        data.append([
            f"OP-{orden.numero_orden}",
            orden.receta.nombre[:20],
            orden.fecha_planificada.strftime('%d/%m/%Y'),
            f"{float(orden.cantidad_planificada):.0f}",
            f"{float(orden.cantidad_producida):.0f}",
            f"{float(orden.merma_real):.0f}",
            f"{float(orden.porcentaje_completado):.0f}%",
            orden.get_estado_display()[:10]
        ])
        
        total_planificado += orden.cantidad_planificada
        total_producido += orden.cantidad_producida
        total_merma += orden.merma_real
    
    # Fila de totales
    data.append([
        'TOTALES',
        '',
        '',
        f"{float(total_planificado):.0f}",
        f"{float(total_producido):.0f}",
        f"{float(total_merma):.0f}",
        '',
        ''
    ])
    
    # Crear tabla
    table = Table(data, colWidths=[60, 120, 60, 40, 40, 40, 45, 60])
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B7355')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F0')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5F5F0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#8B7355')),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="reporte_produccion_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response
