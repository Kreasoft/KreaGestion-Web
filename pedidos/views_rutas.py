"""
Vistas para gestionar Rutas y Hojas de Ruta
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime

from core.decorators import requiere_empresa
from .models_rutas import Ruta, HojaRuta
from .forms_rutas import RutaForm
from .forms_hoja_ruta import HojaRutaForm
from facturacion_electronica.models import DocumentoTributarioElectronico
from django.views.decorators.http import require_http_methods

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required
@requiere_empresa
@permission_required('pedidos.view_ruta', raise_exception=True)
def ruta_list(request):
    """Lista de rutas maestras"""
    rutas = Ruta.objects.filter(empresa=request.empresa).select_related(
        'vehiculo', 'chofer', 'acompanante'
    ).annotate(
        clientes_count=Count('clientes')
    ).order_by('orden_visita', 'codigo')
    
    context = {
        'rutas': rutas,
        'titulo': 'Rutas de Despacho'
    }
    return render(request, 'pedidos/ruta_list.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.add_ruta', raise_exception=True)
def ruta_create(request):
    """Crear nueva ruta"""
    if request.method == 'POST':
        form = RutaForm(request.POST, empresa=request.empresa)
        if form.is_valid():
            ruta = form.save(commit=False)
            ruta.empresa = request.empresa
            ruta.creado_por = request.user
            ruta.save()
            messages.success(request, 'Ruta creada exitosamente.')
            return redirect('pedidos:ruta_detail', pk=ruta.pk)
    else:
        form = RutaForm(empresa=request.empresa)
    
    context = {
        'form': form,
        'titulo': 'Crear Ruta'
    }
    return render(request, 'pedidos/ruta_form.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.view_ruta', raise_exception=True)
def ruta_detail(request, pk):
    """Detalle de ruta"""
    ruta = get_object_or_404(
        Ruta.objects.select_related('vehiculo', 'chofer', 'acompanante', 'empresa'),
        pk=pk, 
        empresa=request.empresa
    )
    
    # Obtener clientes asignados a esta ruta
    clientes = ruta.clientes.filter(empresa=request.empresa).order_by('nombre')
    
    # Obtener hojas de ruta recientes
    hojas_ruta = ruta.hojas_ruta.filter(empresa=request.empresa).order_by('-fecha')[:10]
    
    context = {
        'ruta': ruta,
        'clientes': clientes,
        'hojas_ruta': hojas_ruta,
        'titulo': f'Ruta: {ruta.nombre}'
    }
    return render(request, 'pedidos/ruta_detail.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.change_ruta', raise_exception=True)
def ruta_edit(request, pk):
    """Editar ruta"""
    ruta = get_object_or_404(Ruta, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        form = RutaForm(request.POST, instance=ruta, empresa=request.empresa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ruta actualizada exitosamente.')
            return redirect('pedidos:ruta_detail', pk=ruta.pk)
    else:
        form = RutaForm(instance=ruta, empresa=request.empresa)
    
    context = {
        'form': form,
        'ruta': ruta,
        'titulo': f'Editar Ruta: {ruta.nombre}'
    }
    return render(request, 'pedidos/ruta_form.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.delete_ruta', raise_exception=True)
def ruta_delete(request, pk):
    """Eliminar ruta"""
    ruta = get_object_or_404(Ruta, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        ruta.delete()
        messages.success(request, 'Ruta eliminada exitosamente.')
        return redirect('pedidos:ruta_list')
    
    context = {
        'ruta': ruta,
        'titulo': 'Eliminar Ruta'
    }
    return render(request, 'pedidos/ruta_confirm_delete.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.view_hojaruta', raise_exception=True)
def hoja_ruta_list(request):
    """Lista de hojas de ruta"""
    # Filtros
    estado = request.GET.get('estado', '')
    fecha = request.GET.get('fecha', '')
    ruta_id = request.GET.get('ruta', '')
    search = request.GET.get('search', '')
    
    # Query base
    hojas_ruta = HojaRuta.objects.filter(
        empresa=request.empresa
    ).select_related(
        'ruta', 'vehiculo', 'chofer', 'creado_por'
    ).prefetch_related('facturas')
    
    # Aplicar filtros
    if estado:
        hojas_ruta = hojas_ruta.filter(estado=estado)
    
    if fecha:
        hojas_ruta = hojas_ruta.filter(fecha=fecha)
    
    if ruta_id:
        hojas_ruta = hojas_ruta.filter(ruta_id=ruta_id)
    
    if search:
        hojas_ruta = hojas_ruta.filter(
            Q(numero_ruta__icontains=search) |
            Q(ruta__nombre__icontains=search) |
            Q(chofer__nombre__icontains=search) |
            Q(vehiculo__patente__icontains=search)
        )
    
    # Estadísticas
    total_hojas = hojas_ruta.count()
    pendientes = hojas_ruta.filter(estado='pendiente').count()
    en_ruta = hojas_ruta.filter(estado='en_ruta').count()
    completadas = hojas_ruta.filter(estado='completada').count()
    
    context = {
        'hojas_ruta': hojas_ruta.order_by('-fecha', '-numero_ruta'),
        'rutas': Ruta.objects.filter(empresa=request.empresa, activo=True),
        'estado': estado,
        'fecha': fecha,
        'ruta_id': ruta_id,
        'search': search,
        'total_hojas': total_hojas,
        'pendientes': pendientes,
        'en_ruta': en_ruta,
        'completadas': completadas,
        'titulo': 'Hojas de Ruta'
    }
    return render(request, 'pedidos/hoja_ruta_list.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.view_hojaruta', raise_exception=True)
def hoja_ruta_detail(request, pk):
    """Detalle de hoja de ruta"""
    hoja_ruta = get_object_or_404(
        HojaRuta.objects.select_related('ruta', 'vehiculo', 'chofer', 'creado_por')
        .prefetch_related('facturas'),
        pk=pk,
        empresa=request.empresa
    )
    
    # Obtener facturas con detalles (incluyendo vendedor)
    facturas = hoja_ruta.facturas.all().select_related('venta', 'vendedor', 'venta__vendedor').order_by('folio')
    
    # Calcular totales
    total_neto = sum(f.monto_neto for f in facturas)
    total_iva = sum(f.monto_iva for f in facturas)
    total_total = sum(f.monto_total for f in facturas)
    
    context = {
        'hoja_ruta': hoja_ruta,
        'facturas': facturas,
        'total_neto': total_neto,
        'total_iva': total_iva,
        'total_total': total_total,
        'titulo': f'Hoja de Ruta: {hoja_ruta.numero_ruta}'
    }
    return render(request, 'pedidos/hoja_ruta_detail.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.add_hojaruta', raise_exception=True)
def hoja_ruta_create(request):
    """Crear nueva hoja de ruta manualmente"""
    if request.method == 'POST':
        form = HojaRutaForm(request.POST, empresa=request.empresa)
        if form.is_valid():
            hoja_ruta = form.save(commit=False)
            hoja_ruta.empresa = request.empresa
            hoja_ruta.creado_por = request.user
            
            # Generar número de ruta si no existe
            if not hoja_ruta.numero_ruta:
                fecha_str = hoja_ruta.fecha.strftime('%Y%m%d')
                ruta_codigo = hoja_ruta.ruta.codigo if hoja_ruta.ruta else 'RUTA'
                # Buscar el siguiente número secuencial para este día y ruta
                ultima_hoja = HojaRuta.objects.filter(
                    empresa=request.empresa,
                    fecha=hoja_ruta.fecha,
                    ruta=hoja_ruta.ruta
                ).order_by('-numero_ruta').first()
                
                if ultima_hoja and ultima_hoja.numero_ruta:
                    try:
                        # Extraer el número secuencial del último número
                        partes = ultima_hoja.numero_ruta.split('-')
                        if len(partes) >= 3:
                            ultimo_num = int(partes[-1])
                            nuevo_num = ultimo_num + 1
                        else:
                            nuevo_num = 1
                    except (ValueError, IndexError):
                        nuevo_num = 1
                else:
                    nuevo_num = 1
                
                hoja_ruta.numero_ruta = f"HR-{fecha_str}-{ruta_codigo}-{nuevo_num:03d}"
            
            hoja_ruta.save()
            
            # Buscar y asociar facturas del día y ruta/vehículo que aún no estén en otra hoja de ruta
            if hoja_ruta.fecha:
                # Obtener IDs de facturas que ya están en otras hojas de ruta
                facturas_en_otras_hojas = HojaRuta.objects.filter(
                    empresa=request.empresa
                ).exclude(pk=hoja_ruta.pk).values_list('facturas__id', flat=True)
                
                # Construir filtros de búsqueda
                filtros_facturas = Q(
                    empresa=request.empresa,
                    fecha_emision=hoja_ruta.fecha,
                    tipo_dte__in=['33', '34']  # Solo facturas (33 y 34)
                )
                
                # Si hay ruta, buscar por clientes de la ruta
                if hoja_ruta.ruta:
                    clientes_ruta = hoja_ruta.ruta.clientes.filter(empresa=request.empresa, estado='activo')
                    filtro_ruta = Q(
                        Q(venta__cliente__in=clientes_ruta) |  # Cliente a través de venta directa
                        Q(orden_despacho__cliente__in=clientes_ruta)  # Cliente a través de orden_despacho
                    )
                else:
                    filtro_ruta = Q()
                
                # Si hay vehículo, también buscar por vehículo
                if hoja_ruta.vehiculo:
                    filtro_vehiculo = Q(
                        Q(vehiculo=hoja_ruta.vehiculo) |  # Vehículo directo en DTE
                        Q(venta__vehiculo=hoja_ruta.vehiculo) |  # Vehículo en venta directa
                        Q(orden_despacho__vehiculo=hoja_ruta.vehiculo)  # Vehículo en orden_despacho
                    )
                else:
                    filtro_vehiculo = Q()
                
                # Combinar filtros: ruta O vehículo (si hay ambos, usar ambos)
                if hoja_ruta.ruta and hoja_ruta.vehiculo:
                    filtro_combinado = filtro_ruta | filtro_vehiculo
                elif hoja_ruta.ruta:
                    filtro_combinado = filtro_ruta
                elif hoja_ruta.vehiculo:
                    filtro_combinado = filtro_vehiculo
                else:
                    filtro_combinado = Q()
                
                # Buscar facturas disponibles
                if filtro_combinado:
                    facturas_disponibles = DocumentoTributarioElectronico.objects.filter(
                        filtros_facturas & filtro_combinado
                    ).exclude(id__in=facturas_en_otras_hojas).distinct()
                    
                    if facturas_disponibles.exists():
                        hoja_ruta.facturas.add(*facturas_disponibles)
                        messages.info(request, f'Se asociaron {facturas_disponibles.count()} factura(s) automáticamente.')
            
            messages.success(request, 'Hoja de ruta creada exitosamente.')
            return redirect('pedidos:hoja_ruta_detail', pk=hoja_ruta.pk)
    else:
        form = HojaRutaForm(empresa=request.empresa)
        # Establecer fecha por defecto a hoy
        form.initial['fecha'] = timezone.now().date()
    
    context = {
        'form': form,
        'titulo': 'Crear Hoja de Ruta'
    }
    return render(request, 'pedidos/hoja_ruta_form.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.change_hojaruta', raise_exception=True)
def hoja_ruta_edit(request, pk):
    """Editar hoja de ruta"""
    hoja_ruta = get_object_or_404(HojaRuta, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        form = HojaRutaForm(request.POST, instance=hoja_ruta, empresa=request.empresa)
        if form.is_valid():
            hoja_ruta = form.save()
            
            # Buscar y asociar facturas del día y ruta/vehículo que aún no estén en otra hoja de ruta
            if hoja_ruta.fecha:
                # Obtener IDs de facturas que ya están en otras hojas de ruta
                facturas_en_otras_hojas = HojaRuta.objects.filter(
                    empresa=request.empresa
                ).exclude(pk=hoja_ruta.pk).values_list('facturas__id', flat=True)
                
                # Construir filtros de búsqueda
                filtros_facturas = Q(
                    empresa=request.empresa,
                    fecha_emision=hoja_ruta.fecha,
                    tipo_dte__in=['33', '34']  # Solo facturas (33 y 34)
                )
                
                # Si hay ruta, buscar por clientes de la ruta
                if hoja_ruta.ruta:
                    clientes_ruta = hoja_ruta.ruta.clientes.filter(empresa=request.empresa, estado='activo')
                    filtro_ruta = Q(
                        Q(venta__cliente__in=clientes_ruta) |  # Cliente a través de venta directa
                        Q(orden_despacho__cliente__in=clientes_ruta)  # Cliente a través de orden_despacho
                    )
                else:
                    filtro_ruta = Q()
                
                # Si hay vehículo, también buscar por vehículo
                if hoja_ruta.vehiculo:
                    filtro_vehiculo = Q(
                        Q(vehiculo=hoja_ruta.vehiculo) |  # Vehículo directo en DTE
                        Q(venta__vehiculo=hoja_ruta.vehiculo) |  # Vehículo en venta directa
                        Q(orden_despacho__vehiculo=hoja_ruta.vehiculo)  # Vehículo en orden_despacho
                    )
                else:
                    filtro_vehiculo = Q()
                
                # Combinar filtros: ruta O vehículo (si hay ambos, usar ambos)
                if hoja_ruta.ruta and hoja_ruta.vehiculo:
                    filtro_combinado = filtro_ruta | filtro_vehiculo
                elif hoja_ruta.ruta:
                    filtro_combinado = filtro_ruta
                elif hoja_ruta.vehiculo:
                    filtro_combinado = filtro_vehiculo
                else:
                    filtro_combinado = Q()
                
                # Buscar facturas disponibles
                if filtro_combinado:
                    facturas_disponibles = DocumentoTributarioElectronico.objects.filter(
                        filtros_facturas & filtro_combinado
                    ).exclude(id__in=facturas_en_otras_hojas).distinct()
                    
                    if facturas_disponibles.exists():
                        hoja_ruta.facturas.add(*facturas_disponibles)
                        messages.info(request, f'Se asociaron {facturas_disponibles.count()} factura(s) automáticamente.')
            
            messages.success(request, 'Hoja de ruta actualizada exitosamente.')
            return redirect('pedidos:hoja_ruta_detail', pk=hoja_ruta.pk)
    else:
        form = HojaRutaForm(instance=hoja_ruta, empresa=request.empresa)
    
    context = {
        'form': form,
        'hoja_ruta': hoja_ruta,
        'titulo': f'Editar Hoja de Ruta: {hoja_ruta.numero_ruta}'
    }
    return render(request, 'pedidos/hoja_ruta_form.html', context)


@login_required
@requiere_empresa
@require_http_methods(["GET"])
def ajax_ruta_datos(request, ruta_id):
    """Obtener datos de una ruta (vehículo, chofer, acompañante) para usar como valores por defecto"""
    try:
        ruta = get_object_or_404(Ruta, pk=ruta_id, empresa=request.empresa)
        datos = {
            'vehiculo_id': ruta.vehiculo.pk if ruta.vehiculo else None,
            'chofer_id': ruta.chofer.pk if ruta.chofer else None,
            'acompanante_id': ruta.acompanante.pk if ruta.acompanante else None,
        }
        return JsonResponse(datos)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@requiere_empresa
@require_http_methods(["GET"])
def ajax_vehiculo_chofer(request, vehiculo_id):
    """Obtener el chofer asignado por defecto a un vehículo"""
    try:
        from .models_transporte import Vehiculo
        vehiculo = get_object_or_404(Vehiculo, pk=vehiculo_id, empresa=request.empresa)
        datos = {
            'success': True,
            'chofer_id': vehiculo.chofer.pk if vehiculo.chofer else None,
            'chofer_nombre': vehiculo.chofer.nombre if vehiculo.chofer else None,
        }
        return JsonResponse(datos)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@requiere_empresa
@permission_required('pedidos.view_hojaruta', raise_exception=True)
def hoja_ruta_imprimir(request, pk):
    """Imprimir hoja de ruta"""
    hoja_ruta = get_object_or_404(
        HojaRuta.objects.select_related('ruta', 'vehiculo', 'chofer', 'acompanante', 'empresa', 'creado_por')
        .prefetch_related('facturas'),
        pk=pk,
        empresa=request.empresa
    )
    
    # Obtener facturas con detalles (incluyendo vendedor)
    facturas = hoja_ruta.facturas.all().select_related('venta', 'vendedor', 'venta__vendedor').order_by('folio')
    
    # Calcular totales
    total_neto = sum(f.monto_neto for f in facturas)
    total_iva = sum(f.monto_iva for f in facturas)
    total_total = sum(f.monto_total for f in facturas)
    
    context = {
        'hoja_ruta': hoja_ruta,
        'facturas': facturas,
        'total_neto': total_neto,
        'total_iva': total_iva,
        'total_total': total_total,
        'titulo': f'Hoja de Ruta: {hoja_ruta.numero_ruta}'
    }
    return render(request, 'pedidos/hoja_ruta_imprimir.html', context)


@login_required
@requiere_empresa
@permission_required('pedidos.view_hojaruta', raise_exception=True)
def hoja_ruta_exportar_excel(request, pk):
    """Exportar hoja de ruta a Excel con formato elegante"""
    hoja_ruta = get_object_or_404(
        HojaRuta.objects.select_related('ruta', 'vehiculo', 'chofer', 'acompanante', 'empresa', 'creado_por')
        .prefetch_related('facturas'),
        pk=pk,
        empresa=request.empresa
    )
    
    # Obtener facturas con detalles
    facturas = hoja_ruta.facturas.all().select_related('venta', 'vendedor', 'venta__vendedor').order_by('folio')
    
    # Calcular totales
    total_neto = sum(f.monto_neto for f in facturas)
    total_iva = sum(f.monto_iva for f in facturas)
    total_total = sum(f.monto_total for f in facturas)
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja de Ruta"
    
    # Estilos elegantes
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color="8B7355")
    subtitle_font = Font(bold=True, size=12, color="6F5B44")
    label_font = Font(bold=True, size=10, color="495057")
    normal_font = Font(size=10)
    total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    total_font = Font(bold=True, size=11, color="2C3E50")
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Título principal
    ws.merge_cells('A1:H1')
    ws['A1'] = f'HOJA DE RUTA - {hoja_ruta.numero_ruta}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="F4E4BC", end_color="F4E4BC", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # Información de la empresa
    row = 3
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f'{hoja_ruta.empresa.razon_social}'
    ws[f'A{row}'].font = subtitle_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f'RUT: {hoja_ruta.empresa.rut}'
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2
    
    # Información de la hoja de ruta
    info_data = [
        ('Número de Ruta:', hoja_ruta.numero_ruta),
        ('Fecha:', hoja_ruta.fecha.strftime('%d/%m/%Y') if hoja_ruta.fecha else ''),
        ('Ruta:', f"{hoja_ruta.ruta.codigo} - {hoja_ruta.ruta.nombre}" if hoja_ruta.ruta else ''),
        ('Vehículo:', hoja_ruta.vehiculo.patente if hoja_ruta.vehiculo else ''),
        ('Chofer:', hoja_ruta.chofer.nombre if hoja_ruta.chofer else ''),
        ('Acompañante:', hoja_ruta.acompanante.nombre if hoja_ruta.acompanante else 'No asignado'),
        ('Estado:', hoja_ruta.get_estado_display()),
    ]
    
    for i, (label, value) in enumerate(info_data):
        if i % 2 == 0:
            col = 'A'
            row += 1
        else:
            col = 'E'
        
        ws[f'{col}{row}'] = label
        ws[f'{col}{row}'].font = label_font
        ws[f'{col}{row}'].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        ws[f'{col}{row}'].border = border
        
        # Valor en la siguiente columna
        if col == 'A':
            # Merge de B a D (3 columnas)
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = normal_font
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            ws[f'B{row}'].border = border
        else:
            # Merge de F a H (3 columnas)
            ws.merge_cells(f'F{row}:H{row}')
            ws[f'F{row}'] = value
            ws[f'F{row}'].font = normal_font
            ws[f'F{row}'].alignment = Alignment(horizontal='left')
            ws[f'F{row}'].border = border
    
    row += 2
    
    # Observaciones si existen
    if hoja_ruta.observaciones:
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f'Observaciones: {hoja_ruta.observaciones}'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    
    row += 1
    
    # Encabezado de facturas
    headers = ['N° Factura', 'Fecha', 'Cliente', 'RUT', 'Vendedor', 'Neto', 'IVA', 'Total']
    header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    
    header_row = row  # Guardar la fila del encabezado para freeze_panes
    
    for col, header in zip(header_cols, headers):
        cell = ws[f'{col}{row}']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Datos de facturas
    for factura in facturas:
        # Obtener cliente
        cliente_nombre = 'Sin cliente'
        cliente_rut = ''
        if factura.venta and factura.venta.cliente:
            cliente_nombre = factura.venta.cliente.nombre
            cliente_rut = factura.venta.cliente.rut or ''
        elif factura.orden_despacho.exists():
            primera_venta = factura.orden_despacho.first()
            if primera_venta and primera_venta.cliente:
                cliente_nombre = primera_venta.cliente.nombre
                cliente_rut = primera_venta.cliente.rut or ''
        
        # Obtener vendedor
        vendedor_nombre = ''
        if factura.vendedor:
            vendedor_nombre = factura.vendedor.nombre
        elif factura.venta and factura.venta.vendedor:
            vendedor_nombre = factura.venta.vendedor.nombre
        
        data_row = [
            factura.folio,
            factura.fecha_emision.strftime('%d/%m/%Y') if factura.fecha_emision else '',
            cliente_nombre,
            cliente_rut,
            vendedor_nombre,
            float(factura.monto_neto),
            float(factura.monto_iva),
            float(factura.monto_total),
        ]
        
        for col, value in zip(header_cols, data_row):
            cell = ws[f'{col}{row}']
            cell.value = value
            cell.font = normal_font
            cell.border = border
            
            # Formato numérico para montos
            if col in ['F', 'G', 'H']:
                cell.number_format = '$#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col == 'B':
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top')
        
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Fila de totales
    ws[f'E{row}'] = 'TOTALES:'
    ws[f'E{row}'].font = total_font
    ws[f'E{row}'].fill = total_fill
    ws[f'E{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'].border = border
    
    ws[f'F{row}'] = float(total_neto)
    ws[f'F{row}'].font = total_font
    ws[f'F{row}'].fill = total_fill
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'].border = border
    
    ws[f'G{row}'] = float(total_iva)
    ws[f'G{row}'].font = total_font
    ws[f'G{row}'].fill = total_fill
    ws[f'G{row}'].number_format = '$#,##0'
    ws[f'G{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'].border = border
    
    ws[f'H{row}'] = float(total_total)
    ws[f'H{row}'].font = total_font
    ws[f'H{row}'].fill = total_fill
    ws[f'H{row}'].number_format = '$#,##0'
    ws[f'H{row}'].alignment = Alignment(horizontal='right')
    ws[f'H{row}'].border = border
    
    ws.row_dimensions[row].height = 25
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15  # N° Factura
    ws.column_dimensions['B'].width = 12  # Fecha
    ws.column_dimensions['C'].width = 35  # Cliente
    ws.column_dimensions['D'].width = 15  # RUT
    ws.column_dimensions['E'].width = 20  # Vendedor
    ws.column_dimensions['F'].width = 15  # Neto
    ws.column_dimensions['G'].width = 15  # IVA
    ws.column_dimensions['H'].width = 15  # Total
    
    # Congelar paneles (fijar encabezados) - solo si hay facturas
    if facturas.exists() and header_row > 0:
        # Congelar en la fila siguiente al encabezado
        ws.freeze_panes = f'A{header_row + 1}'
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Hoja_Ruta_{hoja_ruta.numero_ruta}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

