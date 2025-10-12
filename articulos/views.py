from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from decimal import Decimal, InvalidOperation
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfgen import canvas
from datetime import datetime
import os
from .models import Articulo, CategoriaArticulo, UnidadMedida, StockArticulo, ImpuestoEspecifico, ListaPrecio, PrecioArticulo
from inventario.models import Stock
from .forms import ArticuloForm, CategoriaArticuloForm, UnidadMedidaForm, ImpuestoEspecificoForm, ListaPrecioForm, PrecioArticuloForm
from empresas.decorators import requiere_empresa


@requiere_empresa
@login_required
@permission_required('articulos.view_articulo', raise_exception=True)
def articulo_list(request):
    """Lista de artículos"""
    # Filtrar por empresa activa (todos los usuarios, incluyendo superusuarios)
    articulos = Articulo.objects.filter(empresa=request.empresa).order_by('-fecha_creacion')
    
    # Filtros
    search = request.GET.get('search', '')
    categoria_id = request.GET.get('categoria', '')
    activo = request.GET.get('activo', '')
    
    if search:
        articulos = articulos.filter(
            Q(codigo__icontains=search) |
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    if categoria_id:
        articulos = articulos.filter(categoria_id=categoria_id)
    
    if activo != '':
        articulos = articulos.filter(activo=activo == 'true')
    
    # Paginación
    paginator = Paginator(articulos, 25)  # 25 artículos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Agregar stock total a cada artículo
    for articulo in page_obj:
        if articulo.control_stock:
            stock_total = Stock.objects.filter(articulo=articulo).aggregate(
                total=Sum('cantidad')
            )['total']
            articulo.stock_total = stock_total or 0
        else:
            articulo.stock_total = None
    
    # Estadísticas reales (antes de aplicar filtros)
    articulos_totales = Articulo.objects.filter(empresa=request.empresa)
    total_articulos = articulos_totales.count()
    articulos_activos = articulos_totales.filter(activo=True).count()
    articulos_inactivos = articulos_totales.filter(activo=False).count()
    
    # Contar artículos con control de stock
    articulos_con_stock = articulos_totales.filter(control_stock=True).count()
    
    # Categorías para el filtro
    categorias = CategoriaArticulo.objects.filter(empresa=request.empresa, activa=True)
    
    context = {
        'page_obj': page_obj,
        'categorias': categorias,
        'search': search,
        'categoria_id': categoria_id,
        'activo': activo,
        'total_articulos': total_articulos,
        'articulos_activos': articulos_activos,
        'articulos_inactivos': articulos_inactivos,
        'articulos_con_stock': articulos_con_stock,
    }
    
    return render(request, 'articulos/articulo_list.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.view_articulo', raise_exception=True)
def articulo_detail(request, pk):
    """Detalle de un artículo"""
    # Para superusuarios, permitir ver cualquier artículo
    if request.user.is_superuser:
        articulo = get_object_or_404(Articulo, pk=pk)
    else:
        articulo = get_object_or_404(Articulo, pk=pk, empresa=request.empresa)
    
    # Stocks por sucursal
    stocks = StockArticulo.objects.filter(articulo=articulo)
    
    context = {
        'articulo': articulo,
        'stocks': stocks,
    }
    
    return render(request, 'articulos/articulo_detail.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.add_articulo', raise_exception=True)
def articulo_create(request):
    """Crear nuevo artículo"""
    if request.method == 'POST':
        form = ArticuloForm(request.POST, initial={'empresa': request.empresa})
        if form.is_valid():
            try:
                articulo = form.save(commit=False)
                articulo.empresa = request.empresa
                articulo.save()
                messages.success(request, f'Artículo "{articulo.nombre}" creado exitosamente.')
                return redirect('articulos:articulo_list')
            except Exception as e:
                error_msg = str(e)
                if 'UNIQUE constraint failed' in error_msg:
                    if 'codigo_barras' in error_msg:
                        messages.error(request, f'El código de barras "{articulo.codigo_barras}" ya existe en esta empresa. Por favor, use un código de barras diferente.')
                    elif 'codigo' in error_msg:
                        messages.error(request, f'El código "{articulo.codigo}" ya existe en esta empresa. Por favor, use un código diferente.')
                    else:
                        messages.error(request, 'Ya existe un artículo con estos datos en esta empresa.')
                else:
                    messages.error(request, f'Error al guardar el artículo: {error_msg}')
        else:
            messages.error(request, 'Error en el formulario. Revise los datos.')
            # Debug: mostrar errores específicos
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ArticuloForm(initial={'empresa': request.empresa})
    
    context = {
        'form': form,
        'title': 'Crear Artículo',
        'submit_text': 'Crear Artículo',
    }
    
    return render(request, 'articulos/articulo_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.change_articulo', raise_exception=True)
def articulo_update(request, pk):
    """Editar artículo"""
    # Para superusuarios, permitir editar cualquier artículo
    if request.user.is_superuser:
        articulo = get_object_or_404(Articulo, pk=pk)
    else:
        articulo = get_object_or_404(Articulo, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        form = ArticuloForm(request.POST, instance=articulo, initial={'empresa': request.empresa})
        if form.is_valid():
            try:
                articulo = form.save(commit=False)
                articulo.save()
                messages.success(request, f'Artículo "{articulo.nombre}" actualizado exitosamente.')
                return redirect('articulos:articulo_list')
            except Exception as e:
                error_msg = str(e)
                if 'UNIQUE constraint failed' in error_msg:
                    if 'codigo_barras' in error_msg:
                        messages.error(request, f'El código de barras "{articulo.codigo_barras}" ya existe en esta empresa. Por favor, use un código de barras diferente.')
                    elif 'codigo' in error_msg:
                        messages.error(request, f'El código "{articulo.codigo}" ya existe en esta empresa. Por favor, use un código diferente.')
                    else:
                        messages.error(request, 'Ya existe un artículo con estos datos en esta empresa.')
                else:
                    messages.error(request, f'Error al guardar el artículo: {error_msg}')
        else:
            messages.error(request, 'Error en el formulario. Revise los datos.')
            # Debug: mostrar errores específicos
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ArticuloForm(instance=articulo, initial={'empresa': request.empresa})
    
    context = {
        'form': form,
        'articulo': articulo,
        'title': 'Editar Artículo',
        'submit_text': 'Actualizar Artículo',
    }
    
    return render(request, 'articulos/articulo_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.delete_articulo', raise_exception=True)
def articulo_delete(request, pk):
    """Eliminar artículo"""
    # Para superusuarios, permitir eliminar cualquier artículo
    if request.user.is_superuser:
        articulo = get_object_or_404(Articulo, pk=pk)
    else:
        articulo = get_object_or_404(Articulo, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        nombre = articulo.nombre
        articulo.delete()
        messages.success(request, f'Artículo "{nombre}" eliminado exitosamente.')
        return redirect('articulos:articulo_list')
    
    context = {
        'articulo': articulo,
    }
    
    return render(request, 'articulos/articulo_confirm_delete.html', context)


# Vistas para Categorías
@requiere_empresa
@login_required
@permission_required('articulos.view_categoriaarticulo', raise_exception=True)
def categoria_list(request):
    """Lista de categorías"""
    # Filtrar por empresa activa (todos los usuarios, incluyendo superusuarios)
    categorias = CategoriaArticulo.objects.filter(empresa=request.empresa).order_by('nombre')
    
    # Calcular estadísticas
    total_categorias = categorias.count()
    categorias_activas = categorias.filter(activa=True).count()
    categorias_con_iva = categorias.filter(exenta_iva=False).count()
    categorias_exentas = categorias.filter(exenta_iva=True).count()
    
    # Obtener impuestos específicos para el modal
    impuestos_especificos = ImpuestoEspecifico.objects.filter(empresa=request.empresa, activa=True).order_by('nombre')
    
    context = {
        'categorias': categorias,
        'total_categorias': total_categorias,
        'categorias_activas': categorias_activas,
        'categorias_con_iva': categorias_con_iva,
        'categorias_exentas': categorias_exentas,
        'impuestos_especificos': impuestos_especificos,
    }
    
    return render(request, 'articulos/categoria_list.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.add_categoriaarticulo', raise_exception=True)
def categoria_create(request):
    """Crear nueva categoría"""
    if request.method == 'POST':
        form = CategoriaArticuloForm(request.POST, initial={'empresa': request.empresa})
        if form.is_valid():
            categoria = form.save(commit=False)
            categoria.empresa = request.empresa
            categoria.save()
            messages.success(request, f'Categoría "{categoria.nombre}" creada exitosamente.')
            return redirect('articulos:categoria_list')
    else:
        form = CategoriaArticuloForm(initial={'empresa': request.empresa})
    
    context = {
        'form': form,
        'title': 'Crear Categoría',
        'submit_text': 'Crear Categoría',
    }
    
    return render(request, 'articulos/categoria_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.change_categoriaarticulo', raise_exception=True)
def categoria_update(request, pk):
    """Editar categoría"""
    # Para superusuarios, no filtrar por empresa
    if request.user.is_superuser:
        categoria = get_object_or_404(CategoriaArticulo, pk=pk)
    else:
        categoria = get_object_or_404(CategoriaArticulo, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        form = CategoriaArticuloForm(request.POST, instance=categoria, initial={'empresa': request.empresa})
        if form.is_valid():
            form.save()
            messages.success(request, f'Categoría "{categoria.nombre}" actualizada exitosamente.')
            return redirect('articulos:categoria_list')
    else:
        form = CategoriaArticuloForm(instance=categoria, initial={'empresa': request.empresa})
    
    context = {
        'form': form,
        'categoria': categoria,
        'title': 'Editar Categoría',
        'submit_text': 'Actualizar Categoría',
    }
    
    return render(request, 'articulos/categoria_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.delete_categoriaarticulo', raise_exception=True)
def categoria_delete(request, pk):
    """Eliminar categoría"""
    from django.db.models import ProtectedError
    
    # Para superusuarios, no filtrar por empresa
    if request.user.is_superuser:
        categoria = get_object_or_404(CategoriaArticulo, pk=pk)
    else:
        categoria = get_object_or_404(CategoriaArticulo, pk=pk, empresa=request.empresa)
    
    # Verificar si tiene artículos asociados
    articulos_count = categoria.articulo_set.count()
    
    if request.method == 'POST':
        nombre = categoria.nombre
        try:
            categoria.delete()
            messages.success(request, f'Categoría "{nombre}" eliminada exitosamente.')
            return redirect('articulos:categoria_list')
        except ProtectedError:
            messages.error(
                request, 
                f'No se puede eliminar la categoría "{nombre}" porque tiene {articulos_count} artículo(s) asociado(s). '
                f'Primero debe reasignar o eliminar los artículos de esta categoría.'
            )
            return redirect('articulos:categoria_list')
    
    context = {
        'categoria': categoria,
        'articulos_count': articulos_count,
        'title': 'Eliminar Categoría',
    }
    
    return render(request, 'articulos/categoria_confirm_delete.html', context)


# Vistas para Unidades de Medida
@requiere_empresa
@login_required
@permission_required('articulos.view_unidadmedida', raise_exception=True)
def unidad_medida_list(request):
    """Lista de unidades de medida"""
    # Filtrar por empresa activa (todos los usuarios, incluyendo superusuarios)
    unidades = UnidadMedida.objects.filter(empresa=request.empresa).order_by('nombre')
    
    context = {
        'unidades': unidades,
    }
    
    return render(request, 'articulos/unidad_list.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.add_unidadmedida', raise_exception=True)
def unidad_medida_create(request):
    """Crear nueva unidad de medida"""
    if request.method == 'POST':
        form = UnidadMedidaForm(request.POST)
        if form.is_valid():
            unidad = form.save(commit=False)
            unidad.empresa = request.empresa
            unidad.save()
            messages.success(request, f'Unidad "{unidad.nombre}" creada exitosamente.')
            return redirect('articulos:unidad_medida_list')
    else:
        form = UnidadMedidaForm()
    
    context = {
        'form': form,
        'title': 'Crear Unidad de Medida',
        'submit_text': 'Crear Unidad',
    }
    
    return render(request, 'articulos/unidad_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.change_unidadmedida', raise_exception=True)
def unidad_medida_update(request, pk):
    """Editar unidad de medida"""
    # Para superusuarios, no filtrar por empresa
    if request.user.is_superuser:
        unidad = get_object_or_404(UnidadMedida, pk=pk)
    else:
        unidad = get_object_or_404(UnidadMedida, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        form = UnidadMedidaForm(request.POST, instance=unidad)
        if form.is_valid():
            form.save()
            messages.success(request, f'Unidad "{unidad.nombre}" actualizada exitosamente.')
            return redirect('articulos:unidad_medida_list')
    else:
        form = UnidadMedidaForm(instance=unidad)
    
    context = {
        'form': form,
        'unidad': unidad,
        'title': 'Editar Unidad de Medida',
        'submit_text': 'Actualizar Unidad',
    }
    
    return render(request, 'articulos/unidad_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.delete_unidadmedida', raise_exception=True)
def unidad_medida_delete(request, pk):
    """Eliminar unidad de medida"""
    # Para superusuarios, no filtrar por empresa
    if request.user.is_superuser:
        unidad = get_object_or_404(UnidadMedida, pk=pk)
    else:
        unidad = get_object_or_404(UnidadMedida, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        nombre = unidad.nombre
        unidad.delete()
        messages.success(request, f'Unidad "{nombre}" eliminada exitosamente.')
        return redirect('articulos:unidad_medida_list')
    
    context = {
        'unidad': unidad,
        'title': 'Eliminar Unidad de Medida',
    }
    
    return render(request, 'articulos/unidad_confirm_delete.html', context)


# Vistas para Impuestos Específicos
@requiere_empresa
@login_required
@permission_required('articulos.view_impuestoespecifico', raise_exception=True)
def impuesto_especifico_list(request):
    """Lista de impuestos específicos"""
    # Filtrar por empresa activa (todos los usuarios, incluyendo superusuarios)
    impuestos = ImpuestoEspecifico.objects.filter(empresa=request.empresa).order_by('nombre')
    
    context = {
        'impuestos': impuestos,
        'title': 'Impuestos Específicos',
    }
    
    return render(request, 'articulos/impuesto_especifico_list.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.add_impuestoespecifico', raise_exception=True)
def impuesto_especifico_create(request):
    """Crear impuesto específico"""
    if request.method == 'POST':
        form = ImpuestoEspecificoForm(request.POST, initial={'empresa': request.empresa})
        if form.is_valid():
            impuesto = form.save(commit=False)
            impuesto.empresa = request.empresa
            impuesto.save()
            messages.success(request, f'Impuesto "{impuesto.nombre}" creado exitosamente.')
            return redirect('articulos:impuesto_especifico_list')
    else:
        form = ImpuestoEspecificoForm(initial={'empresa': request.empresa})
    
    context = {
        'form': form,
        'title': 'Crear Impuesto Específico',
        'submit_text': 'Crear Impuesto',
    }
    
    return render(request, 'articulos/impuesto_especifico_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.change_impuestoespecifico', raise_exception=True)
def impuesto_especifico_update(request, pk):
    """Editar impuesto específico"""
    # Para superusuarios, no filtrar por empresa
    if request.user.is_superuser:
        impuesto = get_object_or_404(ImpuestoEspecifico, pk=pk)
    else:
        impuesto = get_object_or_404(ImpuestoEspecifico, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        form = ImpuestoEspecificoForm(request.POST, instance=impuesto, initial={'empresa': request.empresa})
        if form.is_valid():
            form.save()
            messages.success(request, f'Impuesto "{impuesto.nombre}" actualizado exitosamente.')
            return redirect('articulos:impuesto_especifico_list')
    else:
        form = ImpuestoEspecificoForm(instance=impuesto, initial={'empresa': request.empresa})
    
    context = {
        'form': form,
        'impuesto': impuesto,
        'title': 'Editar Impuesto Específico',
        'submit_text': 'Actualizar Impuesto',
    }
    
    return render(request, 'articulos/impuesto_especifico_form.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.delete_impuestoespecifico', raise_exception=True)
def impuesto_especifico_delete(request, pk):
    """Eliminar impuesto específico"""
    # Para superusuarios, no filtrar por empresa
    if request.user.is_superuser:
        impuesto = get_object_or_404(ImpuestoEspecifico, pk=pk)
    else:
        impuesto = get_object_or_404(ImpuestoEspecifico, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        nombre = impuesto.nombre
        impuesto.delete()
        messages.success(request, f'Impuesto "{nombre}" eliminado exitosamente.')
        return redirect('articulos:impuesto_especifico_list')
    
    context = {
        'impuesto': impuesto,
        'title': 'Eliminar Impuesto Específico',
    }
    
    return render(request, 'articulos/impuesto_especifico_confirm_delete.html', context)


@csrf_exempt
@login_required
def calcular_precios_articulo(request):
    """Vista AJAX para cálculos bidireccionales de precios"""
    try:
        # Obtener datos del POST
        costo = Decimal(request.POST.get('costo', '0'))
        precio_venta = Decimal(request.POST.get('precio_venta', '0'))
        precio_final = Decimal(request.POST.get('precio_final', '0'))
        margen_porcentaje = Decimal(request.POST.get('margen_porcentaje', '0'))
        impuesto_especifico = Decimal(request.POST.get('impuesto_especifico', '0'))
        modo = request.POST.get('modo', 'margen')  # 'margen', 'precio_venta', 'precio_final'
        
        # Obtener información de la categoría para determinar si está exenta de IVA
        categoria_id = request.POST.get('categoria_id')
        exenta_iva = False
        
        if categoria_id:
            try:
                from .models import CategoriaArticulo
                categoria = CategoriaArticulo.objects.get(pk=categoria_id, empresa=request.user.empresa)
                exenta_iva = categoria.exenta_iva
            except CategoriaArticulo.DoesNotExist:
                pass
        
        # Factores de impuestos
        factor_iva = Decimal('1.00') if exenta_iva else Decimal('1.19')  # IVA 0% o 19%
        factor_especifico = (Decimal('1') + impuesto_especifico / Decimal('100')) if impuesto_especifico > 0 else Decimal('1')
        factor_total = factor_iva * factor_especifico
        
        if modo == 'margen' and costo and margen_porcentaje:
            # Calcular precio neto desde margen
            factor_utilidad = Decimal('1') + (margen_porcentaje / Decimal('100'))
            precio_neto = costo * factor_utilidad
            precio_venta = precio_neto
            precio_final = precio_neto * factor_total
            
        elif modo == 'precio_venta' and precio_venta:
            # Calcular desde precio neto
            precio_neto = precio_venta
            precio_final = precio_neto * factor_total
            if costo > 0:
                margen_porcentaje = ((precio_neto / costo) - Decimal('1')) * Decimal('100')
            
        elif modo == 'precio_final' and precio_final:
            # Calcular precio neto hacia atrás desde precio final
            precio_neto = precio_final / factor_total
            precio_venta = precio_neto
            if costo > 0:
                margen_porcentaje = ((precio_neto / costo) - Decimal('1')) * Decimal('100')
        else:
            return JsonResponse({'error': 'Datos insuficientes para el cálculo'}, status=400)
            
        # Redondear resultados
        precio_neto = round(precio_neto, 2)
        precio_venta = round(precio_venta, 2)
        precio_final = round(precio_final, 2)
        margen_porcentaje = round(margen_porcentaje, 2)
        
        # Calcular impuestos
        iva_porcentaje = Decimal('0.00') if exenta_iva else Decimal('0.19')
        valor_iva = round(precio_neto * iva_porcentaje, 2)
        valor_impuesto_especifico = round(precio_neto * (impuesto_especifico / Decimal('100')), 2)
        
        return JsonResponse({
            'precio_neto': str(precio_neto),
            'precio_venta': str(precio_venta),
            'precio_final': str(precio_final),
            'margen_porcentaje': str(margen_porcentaje),
            'valor_iva': str(valor_iva),
            'valor_impuesto_especifico': str(valor_impuesto_especifico)
        })
        
    except (ValueError, TypeError) as e:
        return JsonResponse({
            'error': f'Error en el cálculo de precios: {str(e)}'
        }, status=400)


@login_required
@requiere_empresa
def categoria_impuesto_especifico(request, categoria_id):
    """Vista AJAX para obtener el impuesto específico de una categoría"""
    try:
        categoria = get_object_or_404(CategoriaArticulo, pk=categoria_id, empresa=request.empresa)
        
        if categoria.impuesto_especifico:
            impuesto_porcentaje = categoria.impuesto_especifico.get_porcentaje_decimal()
            return JsonResponse({
                'success': True,
                'impuesto_especifico_porcentaje': str(impuesto_porcentaje),
                'impuesto_especifico_valor': '0.00',  # Se calculará en el frontend
                'nombre_impuesto': categoria.impuesto_especifico.nombre,
                'exenta_iva': categoria.exenta_iva
            })
        else:
            return JsonResponse({
                'success': True,
                'impuesto_especifico_porcentaje': '0.00',
                'impuesto_especifico_valor': '0.00',
                'nombre_impuesto': 'Ninguno',
                'exenta_iva': categoria.exenta_iva
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@requiere_empresa
def buscar_por_codigo_barras(request):
    """Vista AJAX para buscar artículo por código de barras"""
    try:
        codigo = request.GET.get('codigo', '').strip()
        print(f"=== BÚSQUEDA CÓDIGO DE BARRAS: '{codigo}' (len: {len(codigo)}) ===")
        print(f"Empresa del usuario: {request.empresa.nombre} (ID: {request.empresa.id})")
        
        if not codigo:
            return JsonResponse({
                'success': False,
                'message': 'Código de barras requerido'
            })
        
        # Verificar si existe el código en la base de datos (sin filtros)
        todos = Articulo.objects.filter(codigo_barras__icontains=codigo)
        print(f"Artículos con ese código (sin filtros): {todos.count()}")
        for art in todos:
            print(f"  - ID: {art.id}, Nombre: {art.nombre}, Código: '{art.codigo_barras}', Activo: {art.activo}, Empresa: {art.empresa.id} ({art.empresa.nombre})")
        
        # Buscar artículo por código de barras en la empresa actual
        # Primero intentar coincidencia exacta
        articulo = Articulo.objects.filter(
            codigo_barras=codigo,
            empresa=request.empresa,
            activo=True
        ).first()
        print(f"Coincidencia exacta: {articulo.id if articulo else 'No encontrado'}")
        
        # Si no se encuentra, intentar búsqueda parcial
        if not articulo:
            articulo = Articulo.objects.filter(
                codigo_barras__icontains=codigo,
                empresa=request.empresa,
                activo=True
            ).first()
            print(f"Coincidencia parcial: {articulo.id if articulo else 'No encontrado'}")
        
        if articulo:
            # Obtener información de impuestos de la categoría
            categoria_exenta_iva = articulo.categoria.exenta_iva if articulo.categoria else False
            impuesto_especifico_porcentaje = 0
            if articulo.categoria and articulo.categoria.impuesto_especifico:
                impuesto_especifico_porcentaje = float(articulo.categoria.impuesto_especifico.porcentaje or 0)
            
            return JsonResponse({
                'success': True,
                'articulo': {
                    'id': articulo.id,
                    'nombre': articulo.nombre,
                    'descripcion': articulo.descripcion,
                    'precio_costo': str(articulo.precio_costo),
                    'precio_venta': str(articulo.precio_venta),
                    'precio_final': str(articulo.precio_final),
                    'margen_porcentaje': str(articulo.margen_porcentaje),
                    'categoria_id': articulo.categoria.id if articulo.categoria else None,
                    'categoria_exenta_iva': categoria_exenta_iva,
                    'impuesto_especifico_porcentaje': impuesto_especifico_porcentaje,
                    'unidad_medida_id': articulo.unidad_medida.id if articulo.unidad_medida else None,
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'No se encontró artículo con ese código de barras'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@login_required
def stock_actual(request):
    """Vista para obtener stock actual de múltiples artículos"""
    try:
        print(f"DEBUG - stock_actual: Usuario: {request.user}")
        print(f"DEBUG - stock_actual: Empresa: {getattr(request, 'empresa', None)}")
        
        if request.method != 'POST':
            return JsonResponse({
                'success': False,
                'message': 'Método no permitido'
            })
        
        # Parsear JSON
        data = json.loads(request.body)
        articulo_ids = data.get('articulo_ids', [])
        
        print(f"DEBUG - stock_actual: IDs recibidos: {articulo_ids}")
        
        if not articulo_ids:
            return JsonResponse({
                'success': False,
                'message': 'IDs de artículos requeridos'
            })
        
        # Obtener stock actual de los artículos
        stocks = {}
        
        # Si es superusuario, buscar en todas las empresas
        if request.user.is_superuser:
            for articulo_id in articulo_ids:
                try:
                    # Buscar el artículo sin filtro de empresa
                    articulo = Articulo.objects.get(
                        id=articulo_id,
                        activo=True
                    )
                    
                    # Obtener stock actual
                    stock_actual = articulo.stock_actual
                    stocks[articulo_id] = stock_actual
                    print(f"DEBUG - stock_actual: Artículo {articulo_id} stock: {stock_actual}")
                    
                except Articulo.DoesNotExist:
                    stocks[articulo_id] = 0
                    print(f"DEBUG - stock_actual: Artículo {articulo_id} no encontrado")
        else:
            # Usuario normal, filtrar por empresa
            empresa = getattr(request, 'empresa', None)
            if not empresa:
                return JsonResponse({
                    'success': False,
                    'message': 'Empresa no encontrada'
                })
            
            for articulo_id in articulo_ids:
                try:
                    # Buscar el artículo
                    articulo = Articulo.objects.get(
                        id=articulo_id,
                        empresa=empresa,
                        activo=True
                    )
                    
                    # Obtener stock actual
                    stock_actual = articulo.stock_actual
                    stocks[articulo_id] = stock_actual
                    print(f"DEBUG - stock_actual: Artículo {articulo_id} stock: {stock_actual}")
                    
                except Articulo.DoesNotExist:
                    stocks[articulo_id] = 0
                    print(f"DEBUG - stock_actual: Artículo {articulo_id} no encontrado")
        
        print(f"DEBUG - stock_actual: Stocks finales: {stocks}")
        
        return JsonResponse({
            'success': True,
            'stocks': stocks
        })
        
    except json.JSONDecodeError as e:
        print(f"DEBUG - stock_actual: Error JSON: {e}")
        return JsonResponse({
            'success': False,
            'message': 'JSON inválido'
        })
    except Exception as e:
        print(f"DEBUG - stock_actual: Error general: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# ==================== LISTAS DE PRECIOS ====================

@requiere_empresa
@login_required
def lista_precio_list(request):
    """Lista de listas de precios"""
    listas = ListaPrecio.objects.filter(empresa=request.empresa).order_by('-es_predeterminada', 'nombre')
    
    context = {
        'listas': listas,
        'total_listas': listas.count(),
    }
    
    return render(request, 'articulos/lista_precio_list.html', context)


@requiere_empresa
@login_required
def lista_precio_create(request):
    """Crear lista de precios"""
    if request.method == 'POST':
        form = ListaPrecioForm(request.POST)
        if form.is_valid():
            lista = form.save(commit=False)
            lista.empresa = request.empresa
            lista.save()
            messages.success(request, f'Lista de precios "{lista.nombre}" creada exitosamente.')
            return redirect('articulos:lista_precio_list')
    else:
        form = ListaPrecioForm()
    
    context = {
        'form': form,
        'titulo': 'Nueva Lista de Precios',
    }
    
    return render(request, 'articulos/lista_precio_form.html', context)


@requiere_empresa
@login_required
def lista_precio_update(request, pk):
    """Editar lista de precios"""
    lista = get_object_or_404(ListaPrecio, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        form = ListaPrecioForm(request.POST, instance=lista)
        if form.is_valid():
            form.save()
            messages.success(request, f'Lista de precios "{lista.nombre}" actualizada exitosamente.')
            return redirect('articulos:lista_precio_list')
    else:
        form = ListaPrecioForm(instance=lista)
    
    context = {
        'form': form,
        'lista': lista,
        'titulo': f'Editar Lista: {lista.nombre}',
    }
    
    return render(request, 'articulos/lista_precio_form.html', context)


@requiere_empresa
@login_required
def lista_precio_delete(request, pk):
    """Eliminar lista de precios"""
    lista = get_object_or_404(ListaPrecio, pk=pk, empresa=request.empresa)
    
    if request.method == 'POST':
        nombre = lista.nombre
        lista.delete()
        messages.success(request, f'Lista de precios "{nombre}" eliminada exitosamente.')
        return redirect('articulos:lista_precio_list')
    
    context = {
        'lista': lista,
    }
    
    return render(request, 'articulos/lista_precio_confirm_delete.html', context)


@requiere_empresa
@login_required
def lista_precio_detail(request, pk):
    """Detalle de lista de precios con artículos"""
    lista = get_object_or_404(ListaPrecio, pk=pk, empresa=request.empresa)
    precios_list = PrecioArticulo.objects.filter(lista_precio=lista).select_related('articulo').order_by('articulo__nombre')
    
    # Paginación: 20 artículos por página
    paginator = Paginator(precios_list, 20)
    page_number = request.GET.get('page', 1)
    precios = paginator.get_page(page_number)
    
    context = {
        'lista': lista,
        'precios': precios,
    }
    
    return render(request, 'articulos/lista_precio_detail.html', context)


@requiere_empresa
@login_required
@permission_required('articulos.change_listaprecio', raise_exception=True)
def lista_precio_gestionar_precios(request, pk):
    """Gestionar precios de artículos en una lista"""
    lista = get_object_or_404(ListaPrecio, pk=pk, empresa=request.empresa)
    
    # Manejar aplicación masiva de descuento
    if request.method == 'POST' and 'aplicar_descuento_masivo' in request.POST:
        try:
            descuento = float(request.POST.get('descuento_porcentaje', 0))
            base_calculo = request.POST.get('base_calculo', 'neto')
            categoria_id = request.POST.get('categoria_filtro', '')
            
            # Obtener artículos a actualizar
            articulos_query = Articulo.objects.filter(empresa=request.empresa, activo=True)
            if categoria_id:
                articulos_query = articulos_query.filter(categoria_id=categoria_id)
            
            contador = 0
            for articulo in articulos_query:
                if base_calculo == 'neto':
                    precio_base = float(articulo.precio_venta)
                    nuevo_precio = precio_base * (1 - descuento / 100)
                else:
                    precio_base = float(articulo.precio_final)
                    nuevo_precio = precio_base * (1 - descuento / 100)
                    nuevo_precio = nuevo_precio / 1.19  # Convertir a neto
                
                if nuevo_precio > 0:
                    PrecioArticulo.objects.update_or_create(
                        articulo=articulo,
                        lista_precio=lista,
                        defaults={'precio': Decimal(str(round(nuevo_precio, 2)))}
                    )
                    contador += 1
            
            lista.save(update_fields=['fecha_actualizacion'])
            messages.success(request, f'✅ Descuento aplicado a {contador} artículos')
            return redirect('articulos:lista_precio_gestionar_precios', pk=lista.pk)
        except Exception as e:
            messages.error(request, f'Error al aplicar descuento: {str(e)}')
            return redirect('articulos:lista_precio_gestionar_precios', pk=lista.pk)
    
    if request.method == 'POST':
        # Procesar los precios enviados
        guardados = 0
        eliminados = 0
        
        for key, value in request.POST.items():
            if key.startswith('precio_'):
                articulo_id = key.replace('precio_', '')
                try:
                    articulo = Articulo.objects.get(id=articulo_id, empresa=request.empresa)
                    
                    # Limpiar el valor: eliminar puntos (separadores de miles) y espacios
                    if value and value.strip():
                        value_clean = value.replace('.', '').replace(' ', '').replace(',', '.')
                        precio_value = Decimal(value_clean) if value_clean else None
                    else:
                        precio_value = None
                    
                    if precio_value and precio_value > 0:
                        # Crear o actualizar precio
                        PrecioArticulo.objects.update_or_create(
                            articulo=articulo,
                            lista_precio=lista,
                            defaults={'precio': precio_value}
                        )
                        guardados += 1
                    else:
                        # Eliminar precio si está vacío o es 0
                        deleted_count = PrecioArticulo.objects.filter(
                            articulo=articulo,
                            lista_precio=lista
                        ).delete()[0]
                        if deleted_count > 0:
                            eliminados += 1
                except (Articulo.DoesNotExist, ValueError, InvalidOperation):
                    continue
        
        # Actualizar fecha de modificación de la lista
        lista.save(update_fields=['fecha_actualizacion'])
        
        messages.success(request, f'✅ Precios actualizados: {guardados} guardados, {eliminados} eliminados.')
        return redirect('articulos:lista_precio_detail', pk=lista.pk)
    
    # Obtener todos los artículos con sus precios actuales en esta lista
    articulos = Articulo.objects.filter(empresa=request.empresa, activo=True).select_related('categoria').order_by('nombre')
    
    # Obtener categorías para el filtro
    categorias = CategoriaArticulo.objects.filter(empresa=request.empresa).order_by('nombre')
    
    # Obtener precios actuales de esta lista
    precios_actuales = {}
    for precio in PrecioArticulo.objects.filter(lista_precio=lista).select_related('articulo'):
        # Convertir Decimal a float para evitar problemas en el template
        precios_actuales[precio.articulo_id] = float(precio.precio)
    
    # Agregar el precio actual a cada artículo
    articulos_list = []
    for articulo in articulos:
        precio = precios_actuales.get(articulo.id)
        articulo.precio_en_lista = precio if precio is not None else None
        articulos_list.append(articulo)
    
    context = {
        'lista': lista,
        'articulos': articulos_list,
        'categorias': categorias,
    }
    
    return render(request, 'articulos/lista_precio_gestionar.html', context)


@requiere_empresa
@login_required
def lista_precio_exportar_excel(request, pk):
    """Exportar lista de precios a Excel con formato elegante"""
    lista = get_object_or_404(ListaPrecio, pk=pk, empresa=request.empresa)
    precios = PrecioArticulo.objects.filter(lista_precio=lista).select_related('articulo').order_by('articulo__nombre')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"
    
    # Estilos
    header_font = Font(name='Poppins', size=14, bold=True, color="FFFFFF")
    title_font = Font(name='Poppins', size=18, bold=True, color="6F5B44")
    subtitle_font = Font(name='Poppins', size=11, color="6F5B44")
    data_font = Font(name='Poppins', size=10)
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    title_fill = PatternFill(start_color="F5F0EB", end_color="F5F0EB", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color='D4C4A8'),
        right=Side(style='thin', color='D4C4A8'),
        top=Side(style='thin', color='D4C4A8'),
        bottom=Side(style='thin', color='D4C4A8')
    )
    
    # Título principal
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = f"LISTA DE PRECIOS: {lista.nombre.upper()}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = title_fill
    ws.row_dimensions[1].height = 30
    
    # Información de la lista
    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = f"Empresa: {request.empresa.nombre} | Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # Espacio
    ws.row_dimensions[3].height = 10
    
    # Encabezados
    headers = ['Código', 'Artículo', 'Precio Neto', 'Precio c/IVA', 'Última Actualización']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.row_dimensions[4].height = 25
    
    # Datos
    for row_num, precio in enumerate(precios, 5):
        precio_iva = float(precio.precio) * 1.19
        
        # Código
        cell = ws.cell(row=row_num, column=1)
        cell.value = precio.articulo.codigo
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Artículo
        cell = ws.cell(row=row_num, column=2)
        cell.value = precio.articulo.nombre
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Precio Neto
        cell = ws.cell(row=row_num, column=3)
        cell.value = float(precio.precio)
        cell.font = data_font
        cell.number_format = '$#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        # Precio c/IVA
        cell = ws.cell(row=row_num, column=4)
        cell.value = precio_iva
        cell.font = data_font
        cell.number_format = '$#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        # Fecha
        cell = ws.cell(row=row_num, column=5)
        cell.value = precio.fecha_actualizacion.strftime('%d/%m/%Y %H:%M')
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Lista_Precios_{lista.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


@requiere_empresa
@login_required
def lista_precio_exportar_pdf(request, pk):
    """Exportar lista de precios a PDF con formato elegante"""
    lista = get_object_or_404(ListaPrecio, pk=pk, empresa=request.empresa)
    precios = PrecioArticulo.objects.filter(lista_precio=lista).select_related('articulo').order_by('articulo__nombre')
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Lista_Precios_{lista.nombre}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    # Crear documento
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#6F5B44'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#6F5B44'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Título
    elements.append(Paragraph(f"LISTA DE PRECIOS: {lista.nombre.upper()}", title_style))
    elements.append(Paragraph(f"Empresa: {request.empresa.nombre} | Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    
    # Preparar datos para la tabla
    data = [['Código', 'Artículo', 'Precio Neto', 'Precio c/IVA', 'Actualización']]
    
    for precio in precios:
        precio_iva = float(precio.precio) * 1.19
        data.append([
            precio.articulo.codigo,
            precio.articulo.nombre[:40],  # Limitar longitud
            f"${float(precio.precio):,.0f}".replace(",", "."),
            f"${precio_iva:,.0f}".replace(",", "."),
            precio.fecha_actualizacion.strftime('%d/%m/%Y')
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[1*inch, 3.5*inch, 1.2*inch, 1.2*inch, 1*inch])
    
    # Estilo de tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B7355')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Código
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Artículo
        ('ALIGN', (2, 1), (3, -1), 'RIGHT'),  # Precios
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Fecha
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D4C4A8')),
        
        # Filas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F0EB')])
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    return response


@login_required
@requiere_empresa
def api_listas_precios(request):
    """API para obtener listas de precios activas"""
    try:
        listas = ListaPrecio.objects.filter(
            empresa=request.empresa,
            activa=True
        ).order_by('-es_predeterminada', 'nombre')
        
        listas_data = []
        for lista in listas:
            listas_data.append({
                'id': lista.id,
                'nombre': lista.nombre,
                'descripcion': lista.descripcion,
                'es_predeterminada': lista.es_predeterminada
            })
        
        return JsonResponse({
            'success': True,
            'listas': listas_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)