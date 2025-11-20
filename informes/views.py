from utilidades.utils import clean_id
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Avg, F, Q, DecimalField
from django.db.models.functions import TruncDate, TruncMonth
from datetime import datetime, timedelta
from decimal import Decimal
import json

from empresas.models import Empresa
from ventas.models import Venta, VentaDetalle
from inventario.models import Stock
from tesoreria.models import DocumentoCliente, PagoDocumentoCliente, MovimientoCuentaCorrienteCliente
from caja.models import AperturaCaja
from documentos.models import DocumentoCompra
from core.decorators import requiere_empresa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO

# Importar funciones de exportación adicionales
from .exportaciones import (
    exportar_cuentas_cobrar_excel,
    exportar_pagos_recibidos_excel,
    exportar_stock_bajo_excel,
    exportar_cierres_caja_excel,
    exportar_compras_periodo_excel,
    exportar_utilidad_familias_excel,
    exportar_categorias_excel
)


@login_required
@requiere_empresa
def dashboard_informes(request):
    """Dashboard principal de informes"""
    return render(request, 'informes/dashboard.html')


# ==================== INFORMES DE VENTAS ====================

@login_required
@requiere_empresa
def informe_ventas_periodo(request):
    """Informe de ventas por período"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        # Por defecto: último mes
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Obtener ventas del período
    ventas_query = Venta.objects.filter(
        empresa=request.empresa,
        fecha__gte=fecha_desde,
        fecha__lte=fecha_hasta,
        estado='confirmada'
    ).order_by('fecha')
    
    # Agrupar por fecha manualmente
    ventas_por_fecha = {}
    for venta in ventas_query:
        fecha_key = venta.fecha
        if fecha_key not in ventas_por_fecha:
            ventas_por_fecha[fecha_key] = {
                'fecha_solo': fecha_key,
                'total_ventas': 0,
                'cantidad_ventas': 0
            }
        ventas_por_fecha[fecha_key]['total_ventas'] += float(venta.total)
        ventas_por_fecha[fecha_key]['cantidad_ventas'] += 1
    
    # Convertir a lista ordenada y calcular promedios
    ventas = sorted(ventas_por_fecha.values(), key=lambda x: x['fecha_solo'])
    
    # Calcular totales
    total_general = sum(v['total_ventas'] for v in ventas)
    cantidad_total = sum(v['cantidad_ventas'] for v in ventas)
    ticket_promedio = total_general / cantidad_total if cantidad_total > 0 else 0
    
    # Agregar ticket promedio a cada venta
    for venta in ventas:
        venta['ticket_promedio'] = venta['total_ventas'] / venta['cantidad_ventas'] if venta['cantidad_ventas'] > 0 else 0
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'ventas': ventas,
        'total_general': total_general,
        'cantidad_total': cantidad_total,
        'ticket_promedio': ticket_promedio
    }
    
    return render(request, 'informes/ventas_periodo.html', context)


@login_required
@requiere_empresa
def informe_ventas_vendedor(request):
    """Informe de ventas por vendedor"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Usar vendedor en lugar de usuario_creacion
    ventas_vendedor = Venta.objects.filter(
        empresa=request.empresa,
        fecha__range=[fecha_desde, fecha_hasta],
        estado='confirmada',
        vendedor__isnull=False
    ).values(
        'vendedor__codigo',
        'vendedor__nombre'
    ).annotate(
        total_ventas=Sum('total'),
        cantidad_ventas=Count('id'),
        ticket_promedio=Avg('total')
    ).order_by('-total_ventas')
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'ventas_vendedor': ventas_vendedor
    }
    
    return render(request, 'informes/ventas_vendedor.html', context)


@login_required
@requiere_empresa
def informe_productos_vendidos(request):
    """Informe de productos más vendidos"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    productos = VentaDetalle.objects.filter(
        venta__empresa=request.empresa,
        venta__fecha__range=[fecha_desde, fecha_hasta],
        venta__estado='confirmada'
    ).values(
        'articulo__nombre',
        'articulo__codigo'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        total_vendido=Sum('precio_total')
    ).order_by('-cantidad_vendida')[:50]
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'productos': productos
    }
    
    return render(request, 'informes/productos_vendidos.html', context)


# ==================== INFORMES DE INVENTARIO ====================

@login_required
@requiere_empresa
def informe_stock_actual(request):
    """Informe de stock actual por bodega"""
    from inventario.models import Stock
    from decimal import Decimal
    
    bodega_id = clean_id(clean_id(request.GET.get('bodega_id')))
    
    stocks = Stock.objects.filter(
        bodega__empresa=request.empresa
    ).select_related('articulo', 'bodega')
    
    if bodega_id:
        stocks = stocks.filter(bodega_id=bodega_id)
    
    stocks = stocks.order_by('bodega__nombre', 'articulo__nombre')
    
    # Calcular valorización y agregar a cada stock
    stocks_list = []
    total_valorizacion = Decimal('0')
    articulos_unicos = set()  # Para contar artículos únicos
    
    for stock in stocks:
        # Obtener precio promedio (puede ser None o 0)
        precio_promedio = stock.precio_promedio or Decimal('0')
        cantidad = Decimal(str(stock.cantidad))
        
        # Calcular valorización
        valorizacion = cantidad * precio_promedio
        stock.valorizacion = valorizacion
        total_valorizacion += valorizacion
        
        # Contar artículos únicos (solo si tiene stock > 0)
        if cantidad > 0:
            articulos_unicos.add(stock.articulo.id)
        
        stocks_list.append(stock)
    
    total_articulos = len(articulos_unicos)
    
    context = {
        'stocks': stocks_list,
        'total_valorizacion': total_valorizacion,
        'total_articulos': total_articulos,
    }
    
    return render(request, 'informes/stock_actual.html', context)


@login_required
@requiere_empresa
def informe_stock_bajo(request):
    """Informe de productos con stock bajo"""
    stocks_bajos = Stock.objects.filter(
        bodega__empresa=request.empresa,
        cantidad__lte=F('articulo__stock_minimo')
    ).select_related('articulo', 'bodega').order_by('cantidad')
    
    context = {
        'stocks_bajos': stocks_bajos
    }
    
    return render(request, 'informes/stock_bajo.html', context)


# ==================== INFORMES DE TESORERÍA ====================

@login_required
@requiere_empresa
def informe_cuentas_por_cobrar(request):
    """Informe de cuentas por cobrar (clientes)"""
    documentos = DocumentoCliente.objects.filter(
        empresa=request.empresa,
        saldo_pendiente__gt=0
    ).select_related('cliente').order_by('fecha_vencimiento')
    
    # Clasificar por vencimiento
    hoy = datetime.now().date()
    vencidos = documentos.filter(fecha_vencimiento__lt=hoy)
    por_vencer = documentos.filter(fecha_vencimiento__gte=hoy)
    
    total_por_cobrar = sum(float(d.saldo_pendiente) for d in documentos)
    total_vencido = sum(float(d.saldo_pendiente) for d in vencidos)
    
    context = {
        'documentos': documentos,
        'vencidos': vencidos,
        'por_vencer': por_vencer,
        'total_por_cobrar': total_por_cobrar,
        'total_vencido': total_vencido,
        'today': datetime.now().date()
    }
    
    return render(request, 'informes/cuentas_por_cobrar.html', context)


@login_required
@requiere_empresa
def informe_pagos_recibidos(request):
    """Informe de pagos recibidos por período"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    pagos = PagoDocumentoCliente.objects.filter(
        documento__empresa=request.empresa,
        fecha_pago__date__range=[fecha_desde, fecha_hasta]
    ).select_related('documento', 'documento__cliente').order_by('-fecha_pago')
    
    # Agrupar por forma de pago
    por_forma_pago = pagos.values('forma_pago').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    total_recibido = sum(float(p.monto) for p in pagos)
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'pagos': pagos,
        'por_forma_pago': por_forma_pago,
        'total_recibido': total_recibido
    }
    
    return render(request, 'informes/pagos_recibidos.html', context)


# ==================== INFORMES DE CAJA ====================

@login_required
@requiere_empresa
def informe_cierres_caja(request):
    """Informe de cierres de caja"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    cierres = AperturaCaja.objects.filter(
        caja__empresa=request.empresa,
        estado='cerrada',
        fecha_cierre__gte=fecha_desde,
        fecha_cierre__lte=fecha_hasta
    ).select_related('usuario_cierre', 'caja').order_by('-fecha_cierre')
    
    # Calcular total de ventas
    total_ventas = sum(
        float(c.total_ventas_efectivo or 0) + 
        float(c.total_ventas_tarjeta or 0) + 
        float(c.total_ventas_transferencia or 0) + 
        float(c.total_ventas_cheque or 0) + 
        float(c.total_ventas_credito or 0)
        for c in cierres
    )
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'cierres': cierres,
        'total_ventas': total_ventas
    }
    
    return render(request, 'informes/cierres_caja.html', context)


# ==================== INFORMES DE COMPRAS ====================

@login_required
@requiere_empresa
def informe_compras_periodo(request):
    """Informe de compras por período"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    compras = DocumentoCompra.objects.filter(
        empresa=request.empresa,
        fecha_emision__range=[fecha_desde, fecha_hasta]
    ).select_related('proveedor').order_by('-fecha_emision')
    
    # Agrupar por proveedor
    por_proveedor = compras.values(
        'proveedor__nombre'
    ).annotate(
        total_compras=Sum('total_documento'),
        cantidad_compras=Count('id')
    ).order_by('-total_compras')
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'compras': compras,
        'por_proveedor': por_proveedor,
        'total_compras': sum(float(c.total_documento) for c in compras)
    }
    
    return render(request, 'informes/compras_periodo.html', context)


# ==================== EXPORTACIONES ====================

@login_required
@requiere_empresa
def exportar_ventas_excel(request):
    """Exportar informe de ventas a Excel"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Obtener datos
    ventas_query = Venta.objects.filter(
        empresa=request.empresa,
        fecha__gte=fecha_desde,
        fecha__lte=fecha_hasta,
        estado='confirmada'
    ).order_by('fecha')
    
    # Agrupar por fecha
    ventas_por_fecha = {}
    for venta in ventas_query:
        fecha_key = venta.fecha
        if fecha_key not in ventas_por_fecha:
            ventas_por_fecha[fecha_key] = {
                'fecha_solo': fecha_key,
                'total_ventas': 0,
                'cantidad_ventas': 0
            }
        ventas_por_fecha[fecha_key]['total_ventas'] += float(venta.total)
        ventas_por_fecha[fecha_key]['cantidad_ventas'] += 1
    
    ventas = sorted(ventas_por_fecha.values(), key=lambda x: x['fecha_solo'])
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Período"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = f"Informe de Ventas - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    
    # Headers
    headers = ['Fecha', 'Cantidad Ventas', 'Total Ventas', 'Ticket Promedio']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 5
    for venta in ventas:
        ws.cell(row=row, column=1, value=venta['fecha_solo'].strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=2, value=venta['cantidad_ventas']).border = border
        ws.cell(row=row, column=3, value=venta['total_ventas']).border = border
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ticket_prom = venta['total_ventas'] / venta['cantidad_ventas'] if venta['cantidad_ventas'] > 0 else 0
        ws.cell(row=row, column=4, value=ticket_prom).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0'
        row += 1
    
    # Totales
    total_general = sum(v['total_ventas'] for v in ventas)
    cantidad_total = sum(v['cantidad_ventas'] for v in ventas)
    ticket_promedio = total_general / cantidad_total if cantidad_total > 0 else 0
    
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=cantidad_total).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_general).font = Font(bold=True)
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=4, value=ticket_promedio).font = Font(bold=True)
    ws.cell(row=row, column=4).number_format = '$#,##0'
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ventas_{fecha_desde}_{fecha_hasta}.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_productos_excel(request):
    """Exportar productos más vendidos a Excel"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    productos = VentaDetalle.objects.filter(
        venta__empresa=request.empresa,
        venta__fecha__range=[fecha_desde, fecha_hasta],
        venta__estado='confirmada'
    ).values(
        'articulo__nombre',
        'articulo__codigo'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        total_vendido=Sum('precio_total')
    ).order_by('-cantidad_vendida')[:50]
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Más Vendidos"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = f"Productos Más Vendidos - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    
    # Headers
    headers = ['#', 'Código', 'Producto', 'Cantidad', 'Total Vendido']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 5
    for idx, producto in enumerate(productos, start=1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=producto['articulo__codigo']).border = border
        ws.cell(row=row, column=3, value=producto['articulo__nombre']).border = border
        ws.cell(row=row, column=4, value=float(producto['cantidad_vendida'])).border = border
        ws.cell(row=row, column=5, value=float(producto['total_vendido'])).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=productos_vendidos_{fecha_desde}_{fecha_hasta}.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_stock_excel(request):
    """Exportar stock actual a Excel"""
    bodega_id = clean_id(clean_id(request.GET.get('bodega_id')))
    
    stocks = Stock.objects.filter(
        bodega__empresa=request.empresa
    ).select_related('articulo', 'bodega')
    
    if bodega_id:
        stocks = stocks.filter(bodega_id=bodega_id)
    
    stocks = stocks.order_by('bodega__nombre', 'articulo__nombre')
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Actual"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = f"Stock Actual - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ['Bodega', 'Código', 'Artículo', 'Stock', 'Precio Promedio', 'Valorización']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 4
    total_valorizacion = 0
    for stock in stocks:
        valorizacion = float(stock.cantidad) * float(stock.precio_promedio)
        total_valorizacion += valorizacion
        
        ws.cell(row=row, column=1, value=stock.bodega.nombre).border = border
        ws.cell(row=row, column=2, value=stock.articulo.codigo).border = border
        ws.cell(row=row, column=3, value=stock.articulo.nombre).border = border
        ws.cell(row=row, column=4, value=float(stock.cantidad)).border = border
        ws.cell(row=row, column=5, value=float(stock.precio_promedio)).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6, value=valorizacion).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        row += 1
    
    # Total
    ws.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_valorizacion).font = Font(bold=True)
    ws.cell(row=row, column=6).number_format = '$#,##0'
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=stock_actual.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_clientes_excel(request):
    """Exportar clientes a Excel"""
    from clientes.models import Cliente
    
    clientes = Cliente.objects.filter(
        empresa=request.empresa
    ).order_by('nombre')
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = f"Listado de Clientes - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ['RUT', 'Nombre', 'Giro', 'Dirección', 'Comuna', 'Ciudad', 'Región', 'Teléfono', 'Email']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 4
    for cliente in clientes:
        ws.cell(row=row, column=1, value=cliente.rut).border = border
        ws.cell(row=row, column=2, value=cliente.nombre).border = border
        ws.cell(row=row, column=3, value=cliente.giro or '').border = border
        ws.cell(row=row, column=4, value=cliente.direccion or '').border = border
        ws.cell(row=row, column=5, value=cliente.comuna or '').border = border
        ws.cell(row=row, column=6, value=cliente.ciudad or '').border = border
        ws.cell(row=row, column=7, value=cliente.region or '').border = border
        ws.cell(row=row, column=8, value=cliente.telefono or '').border = border
        ws.cell(row=row, column=9, value=cliente.email or '').border = border
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_proveedores_excel(request):
    """Exportar proveedores a Excel"""
    from proveedores.models import Proveedor
    
    proveedores = Proveedor.objects.filter(
        empresa=request.empresa
    ).order_by('nombre')
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = f"Listado de Proveedores - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ['RUT', 'Nombre', 'Giro', 'Dirección', 'Comuna', 'Ciudad', 'Teléfono', 'Email']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 4
    for proveedor in proveedores:
        ws.cell(row=row, column=1, value=proveedor.rut).border = border
        ws.cell(row=row, column=2, value=proveedor.nombre).border = border
        ws.cell(row=row, column=3, value=proveedor.giro or '').border = border
        ws.cell(row=row, column=4, value=proveedor.direccion or '').border = border
        ws.cell(row=row, column=5, value=proveedor.comuna or '').border = border
        ws.cell(row=row, column=6, value=proveedor.ciudad or '').border = border
        ws.cell(row=row, column=7, value=proveedor.telefono or '').border = border
        ws.cell(row=row, column=8, value=proveedor.email or '').border = border
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_articulos_excel(request):
    """Exportar artículos a Excel"""
    from articulos.models import Articulo
    
    articulos = Articulo.objects.filter(
        empresa=request.empresa
    ).select_related('categoria', 'unidad_medida').order_by('codigo')
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Artículos"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = f"Listado de Artículos - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ['Código', 'Nombre', 'Descripción', 'Categoría', 'Unidad', 'Precio Costo', 'Precio Venta', 'Precio Final', 'Margen %', 'Código Barras']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 4
    for articulo in articulos:
        ws.cell(row=row, column=1, value=articulo.codigo).border = border
        ws.cell(row=row, column=2, value=articulo.nombre).border = border
        ws.cell(row=row, column=3, value=articulo.descripcion or '').border = border
        ws.cell(row=row, column=4, value=articulo.categoria.nombre if articulo.categoria else '').border = border
        ws.cell(row=row, column=5, value=articulo.unidad_medida.nombre if articulo.unidad_medida else '').border = border
        ws.cell(row=row, column=6, value=float(articulo.precio_costo)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=7, value=float(articulo.precio_venta)).border = border
        ws.cell(row=row, column=7).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=float(articulo.precio_final)).border = border
        ws.cell(row=row, column=8).number_format = '$#,##0'
        ws.cell(row=row, column=9, value=float(articulo.margen_porcentaje)).border = border
        ws.cell(row=row, column=9).number_format = '0.00"%"'
        ws.cell(row=row, column=10, value=articulo.codigo_barras or '').border = border
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 18
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=articulos.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def informe_utilidad_familias(request):
    """Informe de costos, ventas y utilidad por familia"""
    from articulos.models import CategoriaArticulo
    
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    familia_id = request.GET.get('familia')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Query base de ventas por familia
    ventas_query = VentaDetalle.objects.filter(
        venta__empresa=request.empresa,
        venta__fecha__range=[fecha_desde, fecha_hasta],
        venta__estado='confirmada',
        articulo__categoria__isnull=False
    ).values(
        'articulo__categoria__id',
        'articulo__categoria__nombre'
    ).annotate(
        total_ventas=Sum(F('cantidad') * F('precio_unitario'), output_field=DecimalField()),
        total_costo=Sum(F('cantidad') * F('articulo__precio_costo'), output_field=DecimalField()),
        cantidad_vendida=Sum('cantidad'),
        num_ventas=Count('venta', distinct=True)
    )
    
    # Filtrar por familia si se especifica
    if familia_id:
        ventas_query = ventas_query.filter(articulo__categoria__id=familia_id)
    
    ventas_query = ventas_query.order_by('-total_ventas')
    
    # Calcular utilidad para cada familia
    resultados = []
    total_ventas_general = 0
    total_costo_general = 0
    
    for item in ventas_query:
        ventas = float(item['total_ventas'] or 0)
        costo = float(item['total_costo'] or 0)
        utilidad = ventas - costo
        margen = (utilidad / ventas * 100) if ventas > 0 else 0
        
        resultados.append({
            'familia_id': item['articulo__categoria__id'],
            'familia_nombre': item['articulo__categoria__nombre'],
            'total_ventas': ventas,
            'total_costo': costo,
            'utilidad': utilidad,
            'margen_porcentaje': margen,
            'cantidad_vendida': float(item['cantidad_vendida']),
            'num_ventas': item['num_ventas']
        })
        
        total_ventas_general += ventas
        total_costo_general += costo
    
    utilidad_general = total_ventas_general - total_costo_general
    margen_general = (utilidad_general / total_ventas_general * 100) if total_ventas_general > 0 else 0
    
    # Obtener todas las familias para el filtro
    familias = CategoriaArticulo.objects.filter(empresa=request.empresa).order_by('nombre')
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'familia_id': familia_id,
        'familias': familias,
        'resultados': resultados,
        'total_ventas_general': total_ventas_general,
        'total_costo_general': total_costo_general,
        'utilidad_general': utilidad_general,
        'margen_general': margen_general
    }
    
    return render(request, 'informes/utilidad_familias.html', context)


@login_required
@requiere_empresa
def informe_utilidad_familias_detalle(request, familia_id):
    """API para obtener detalle de artículos por familia"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Query de artículos de la familia
    articulos_query = VentaDetalle.objects.filter(
        venta__empresa=request.empresa,
        venta__fecha__range=[fecha_desde, fecha_hasta],
        venta__estado='confirmada',
        articulo__categoria__id=familia_id
    ).values(
        'articulo__id',
        'articulo__codigo',
        'articulo__nombre'
    ).annotate(
        total_ventas=Sum(F('cantidad') * F('precio_unitario'), output_field=DecimalField()),
        total_costo=Sum(F('cantidad') * F('articulo__precio_costo'), output_field=DecimalField()),
        cantidad_vendida=Sum('cantidad')
    ).order_by('-total_ventas')
    
    # Preparar datos para JSON
    articulos = []
    for item in articulos_query:
        ventas = float(item['total_ventas'] or 0)
        costo = float(item['total_costo'] or 0)
        utilidad = ventas - costo
        margen = (utilidad / ventas * 100) if ventas > 0 else 0
        
        articulos.append({
            'codigo': item['articulo__codigo'],
            'nombre': item['articulo__nombre'],
            'cantidad': float(item['cantidad_vendida']),
            'ventas': ventas,
            'costos': costo,
            'utilidad': utilidad,
            'margen': margen
        })
    
    return JsonResponse({
        'success': True,
        'articulos': articulos
    })


@login_required
@requiere_empresa
def exportar_ventas_vendedor_excel(request):
    """Exportar ventas por vendedor a Excel"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    ventas_vendedor = Venta.objects.filter(
        empresa=request.empresa,
        fecha__range=[fecha_desde, fecha_hasta],
        estado='confirmada',
        vendedor__isnull=False
    ).values(
        'vendedor__codigo',
        'vendedor__nombre'
    ).annotate(
        total_ventas=Sum('total'),
        cantidad_ventas=Count('id'),
        ticket_promedio=Avg('total')
    ).order_by('-total_ventas')
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Vendedor"
    
    # Estilos
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = f"Ventas por Vendedor - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    
    # Headers
    headers = ['Código', 'Vendedor', 'Cant. Ventas', 'Total Vendido', 'Ticket Promedio']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 5
    for vendedor in ventas_vendedor:
        ws.cell(row=row, column=1, value=vendedor['vendedor__codigo']).border = border
        ws.cell(row=row, column=2, value=vendedor['vendedor__nombre']).border = border
        ws.cell(row=row, column=3, value=vendedor['cantidad_ventas']).border = border
        ws.cell(row=row, column=4, value=float(vendedor['total_ventas'])).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=5, value=float(vendedor['ticket_promedio'])).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ventas_vendedor_{fecha_desde}_{fecha_hasta}.xlsx'
    wb.save(response)
    return response
