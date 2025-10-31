"""
Funciones de exportación para informes
"""
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, F, Q
from datetime import datetime, timedelta
from decimal import Decimal

from core.decorators import requiere_empresa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ventas.models import Venta
from inventario.models import Stock
from tesoreria.models import DocumentoCliente, PagoDocumentoCliente
from caja.models import AperturaCaja
from documentos.models import DocumentoCompra


@login_required
@requiere_empresa
def exportar_cuentas_cobrar_excel(request):
    """Exportar cuentas por cobrar a Excel"""
    documentos = DocumentoCliente.objects.filter(
        empresa=request.empresa,
        estado__in=['pendiente', 'parcial']
    ).select_related('cliente').order_by('fecha_vencimiento')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas por Cobrar"
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = f"Cuentas por Cobrar - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['N° Documento', 'Cliente', 'Fecha Emisión', 'Fecha Vencimiento', 'Total', 'Pagado', 'Saldo', 'Estado']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row = 4
    for doc in documentos:
        ws.cell(row=row, column=1, value=doc.numero_documento).border = border
        ws.cell(row=row, column=2, value=doc.cliente.nombre if doc.cliente else 'Sin cliente').border = border
        ws.cell(row=row, column=3, value=doc.fecha_emision.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=4, value=doc.fecha_vencimiento.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=5, value=float(doc.total)).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6, value=float(doc.monto_pagado)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=7, value=float(doc.saldo)).border = border
        ws.cell(row=row, column=7).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=doc.get_estado_display()).border = border
        row += 1
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cuentas_por_cobrar.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_pagos_recibidos_excel(request):
    """Exportar pagos recibidos a Excel"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    pagos = PagoDocumentoCliente.objects.filter(
        empresa=request.empresa,
        fecha_pago__range=[fecha_desde, fecha_hasta]
    ).select_related('documento__cliente', 'forma_pago').order_by('-fecha_pago')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos Recibidos"
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = f"Pagos Recibidos - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    
    headers = ['Fecha', 'N° Documento', 'Cliente', 'Monto', 'Forma de Pago']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row = 5
    for pago in pagos:
        ws.cell(row=row, column=1, value=pago.fecha_pago.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=2, value=pago.documento.numero_documento).border = border
        ws.cell(row=row, column=3, value=pago.documento.cliente.nombre if pago.documento.cliente else 'Sin cliente').border = border
        ws.cell(row=row, column=4, value=float(pago.monto)).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=5, value=pago.forma_pago.nombre if pago.forma_pago else 'Sin especificar').border = border
        row += 1
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=pagos_recibidos_{fecha_desde}_{fecha_hasta}.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_stock_bajo_excel(request):
    """Exportar stock bajo a Excel"""
    stocks_bajo = Stock.objects.filter(
        empresa=request.empresa,
        cantidad__lte=F('stock_minimo')
    ).select_related('articulo', 'bodega').order_by('cantidad')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Bajo"
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = f"Productos con Stock Bajo - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Código', 'Artículo', 'Bodega', 'Stock Actual', 'Stock Mínimo', 'Diferencia']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row = 4
    for stock in stocks_bajo:
        diferencia = float(stock.stock_minimo - stock.cantidad)
        ws.cell(row=row, column=1, value=stock.articulo.codigo).border = border
        ws.cell(row=row, column=2, value=stock.articulo.nombre).border = border
        ws.cell(row=row, column=3, value=stock.bodega.nombre).border = border
        ws.cell(row=row, column=4, value=float(stock.cantidad)).border = border
        ws.cell(row=row, column=5, value=float(stock.stock_minimo)).border = border
        ws.cell(row=row, column=6, value=diferencia).border = border
        row += 1
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock_bajo.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_cierres_caja_excel(request):
    """Exportar cierres de caja a Excel"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    cierres = AperturaCaja.objects.filter(
        caja__empresa=request.empresa,
        fecha_apertura__range=[fecha_desde, fecha_hasta],
        estado='cerrada'
    ).select_related('caja', 'usuario_apertura', 'usuario_cierre').order_by('-fecha_apertura')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cierres de Caja"
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = f"Cierres de Caja - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    
    headers = ['Caja', 'Fecha Apertura', 'Fecha Cierre', 'Monto Inicial', 'Total Ventas', 'Monto Final', 'Diferencia']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row = 5
    for cierre in cierres:
        ws.cell(row=row, column=1, value=cierre.caja.nombre).border = border
        ws.cell(row=row, column=2, value=cierre.fecha_apertura.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row, column=3, value=cierre.fecha_cierre.strftime('%d/%m/%Y %H:%M') if cierre.fecha_cierre else '').border = border
        ws.cell(row=row, column=4, value=float(cierre.monto_inicial)).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=5, value=float(cierre.total_ventas)).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6, value=float(cierre.monto_cierre)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=7, value=float(cierre.diferencia)).border = border
        ws.cell(row=row, column=7).number_format = '$#,##0'
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=cierres_caja_{fecha_desde}_{fecha_hasta}.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_utilidad_familias_excel(request):
    """Exportar utilidad por familias a Excel"""
    from articulos.models import CategoriaArticulo
    
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    familia_id = request.GET.get('familia')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Query de ventas por familia
    from ventas.models import VentaDetalle
    from django.db.models import F
    
    from django.db.models import DecimalField as DField
    
    ventas_query = VentaDetalle.objects.filter(
        venta__empresa=request.empresa,
        venta__fecha__range=[fecha_desde, fecha_hasta],
        venta__estado='confirmada',
        articulo__categoria__isnull=False
    ).values(
        'articulo__categoria__nombre'
    ).annotate(
        total_ventas=Sum(F('cantidad') * F('precio_unitario'), output_field=DField()),
        total_costo=Sum(F('cantidad') * F('articulo__precio_costo'), output_field=DField()),
        cantidad_vendida=Sum('cantidad'),
        num_ventas=Count('venta', distinct=True)
    )
    
    if familia_id:
        ventas_query = ventas_query.filter(articulo__categoria__id=familia_id)
    
    ventas_query = ventas_query.order_by('-total_ventas')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilidad por Familias"
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = f"Utilidad por Familias - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    
    headers = ['Familia', 'Cant. Vendida', 'N° Ventas', 'Total Ventas', 'Total Costos', 'Utilidad', 'Margen %']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row = 5
    total_ventas = 0
    total_costos = 0
    
    for item in ventas_query:
        ventas = float(item['total_ventas'] or 0)
        costo = float(item['total_costo'] or 0)
        utilidad = ventas - costo
        margen = (utilidad / ventas * 100) if ventas > 0 else 0
        
        ws.cell(row=row, column=1, value=item['articulo__categoria__nombre']).border = border
        ws.cell(row=row, column=2, value=float(item['cantidad_vendida'])).border = border
        ws.cell(row=row, column=3, value=item['num_ventas']).border = border
        ws.cell(row=row, column=4, value=ventas).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=5, value=costo).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6, value=utilidad).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=7, value=margen).border = border
        ws.cell(row=row, column=7).number_format = '0.0"%"'
        
        total_ventas += ventas
        total_costos += costo
        row += 1
    
    # Totales
    if row > 5:
        ws.cell(row=row, column=1, value='TOTALES').border = border
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_ventas).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_costos).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_ventas - total_costos).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=6).font = Font(bold=True)
        margen_total = ((total_ventas - total_costos) / total_ventas * 100) if total_ventas > 0 else 0
        ws.cell(row=row, column=7, value=margen_total).border = border
        ws.cell(row=row, column=7).number_format = '0.0"%"'
        ws.cell(row=row, column=7).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=utilidad_familias_{fecha_desde}_{fecha_hasta}.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_compras_periodo_excel(request):
    """Exportar compras por período a Excel"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde or not fecha_hasta:
        fecha_hasta = datetime.now().date()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    compras = DocumentoCompra.objects.filter(
        empresa=request.empresa,
        fecha__range=[fecha_desde, fecha_hasta]
    ).select_related('proveedor').order_by('-fecha')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = f"Compras por Período - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    
    headers = ['Fecha', 'N° Documento', 'Proveedor', 'Tipo', 'Neto', 'IVA', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row = 5
    for compra in compras:
        ws.cell(row=row, column=1, value=compra.fecha.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=2, value=compra.numero_documento).border = border
        ws.cell(row=row, column=3, value=compra.proveedor.nombre if compra.proveedor else 'Sin proveedor').border = border
        ws.cell(row=row, column=4, value=compra.get_tipo_documento_display()).border = border
        ws.cell(row=row, column=5, value=float(compra.neto)).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6, value=float(compra.iva)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=7, value=float(compra.total_documento)).border = border
        ws.cell(row=row, column=7).number_format = '$#,##0'
        row += 1
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=compras_{fecha_desde}_{fecha_hasta}.xlsx'
    wb.save(response)
    return response


@login_required
@requiere_empresa
def exportar_categorias_excel(request):
    """Exportar categorías a Excel"""
    from articulos.models import Categoria
    
    categorias = Categoria.objects.filter(
        empresa=request.empresa
    ).annotate(
        total_articulos=Count('articulos')
    ).order_by('nombre')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    
    header_fill = PatternFill(start_color="8B7355", end_color="8B7355", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = f"Categorías - {request.empresa.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    headers = ['Código', 'Nombre', 'Descripción', 'Total Artículos', 'Estado']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row = 5
    for categoria in categorias:
        ws.cell(row=row, column=1, value=categoria.codigo).border = border
        ws.cell(row=row, column=2, value=categoria.nombre).border = border
        ws.cell(row=row, column=3, value=categoria.descripcion or '').border = border
        ws.cell(row=row, column=4, value=categoria.total_articulos).border = border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value='Activa' if categoria.activo else 'Inactiva').border = border
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        row += 1
    
    # Totales
    row += 1
    ws.cell(row=row, column=1, value='TOTAL CATEGORÍAS:').font = Font(bold=True)
    ws.cell(row=row, column=2, value=categorias.count()).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=categorias_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response
